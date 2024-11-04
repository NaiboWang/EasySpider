# 工具库
import csv
import datetime
import json
import os
import sys
import re
import time
import uuid
from bs4 import BeautifulSoup
# import keyboard
from openpyxl import Workbook, load_workbook
# import pandas as pd
# import xlsxwriter
import requests
from urllib.parse import urlparse
import pymysql
from lxml import etree
import smtplib
from email.mime.text import MIMEText
from email.header import Header
import urllib.request
import base64

def send_email(config):
    """
    发送邮件的函数。

    :param config: 包含邮件配置信息的字典。
    """
    # 校验配置信息是否完整
    # required_keys = ["host", "port", "username", "password", "from", "to", "subject", "content"]
    # missing_keys = [key for key in required_keys if key not in config]
    # if missing_keys:
    #     raise ValueError(f"邮件配置缺少必要的键: {', '.join(missing_keys)}")
    try:
        print("正在发送邮件到：" + config['to'])
        message = MIMEText(config['content'], 'plain', 'utf-8')
        message['From'] = Header(f"{config['username'].split('@')[0]} <{config['username']}>")
        to_name_list = []
        for address in config['to'].split(','):
            address = address.strip()
            name = address.split('@')[0]
            to_name_list.append(f"{name} <{address}>")
        to_name_list = ', '.join(to_name_list)
        message['To'] = Header(to_name_list)
        message['Subject'] = Header(config['subject'], 'utf-8')
        # 使用SSL加密方式连接邮件服务器
        smtp_server = smtplib.SMTP_SSL(config['host'], config['port'])
        smtp_server.login(config['username'], config['password'])
        to_address_list = config['to'].split(',')
        smtp_server.sendmail(config['username'], to_address_list, message.as_string())
        print("邮件发送成功|Email sent successfully")
    except Exception as e:
        print(f"无法发送邮件，发生错误：{e}")
        print(f"Failed to send email, error: {e}")
    finally:
        try:
            smtp_server.quit()
        except:
            pass
  
def rename_downloaded_file(stop_event):
    """
    监控E:\report目录,重命名下载完成的文件为 'ivt-result' 并保持原文件扩展名。
    确保下载目录中始终只有一个文件。
    
    :param stop_event: 用于停止监控的事件
    """
    download_dir = r"E:\report"
    
    # 确保目录存在
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)
        
    while not stop_event.is_set():
        try:
            files = os.listdir(download_dir)
            for file in files:
                full_path = os.path.join(download_dir, file)
                
                # 跳过部分临时文件和已重命名的文件
                if full_path.endswith('.crdownload') or full_path.endswith('.htm') or \
                   full_path.endswith('.html') or full_path.startswith('ivt-result'):
                    continue

                # 获取文件扩展名
                _, ext = os.path.splitext(file)
                ext = ext.lower()

                # 构建新的文件名
                new_name = f"ivt-result{ext}"
                new_path = os.path.join(download_dir, new_name)

                # 如果已存在旧文件,删除它
                if os.path.exists(new_path):
                    try:
                        os.remove(new_path)
                        print(f"已删除旧文件: {new_path}")
                    except Exception as e:
                        print(f"无法删除旧文件 {new_path},错误: {e}")
                        continue  # 如果无法删除旧文件,则跳过当前文件

                # 重命名新文件
                try:
                    os.rename(full_path, new_path)
                    print(f"文件已重命名为: {new_path}")
                except Exception as e:
                    print(f"文件重命名失败: {full_path} -> {new_path},错误: {e}")

        except Exception as e:
            print(f"监控下载目录时出错: {e}")

        time.sleep(1)  # 每秒检查一次
    print("下载文件重命名监控已停止。")

def is_valid_url(url):
    try:
        result = urlparse(url)
        return all([result.scheme, result.netloc])
    except ValueError:
        return False


def lowercase_tags_in_xpath(xpath):
    return re.sub(r"([A-Z]+)(?=[\[\]//]|$)", lambda x: x.group(0).lower(), xpath)

# 提取HTML中的文本内容
def extract_text_from_html(html_content):
    soup = BeautifulSoup(html_content, 'lxml') # 使用lxml作为解析器
    for script in soup(["script", "style"]): # 去除脚本和样式内容
        script.extract()
    for p_tag in soup.find_all("p"):
        p_tag.append(soup.new_tag("br")) # 在每个p标签后添加br标签
        p_tag.append("\n") # 在每个p标签后添加换行符
    text = soup.get_text()
    return text

# 将文本按照行分割并去除额外空白
def split_text_by_lines(text):
    lines = text.splitlines()
    lines = [line.strip() for line in lines if line.strip()]  # 去除空行和首尾空格
    return "\n".join(lines)

def on_press_creator(press_time, event):
    def on_press(key):
        try:
            if key.char == press_time["pause_key"]:
                if press_time["is_pressed"] == False:  # 没按下p键时，记录按下p键的时间
                    press_time["duration"] = time.time()
                    press_time["is_pressed"] = True
                else:  # 按下p键时，判断按下p键的时间是否超过2.5秒
                    duration = time.time() - press_time["duration"]
                    if duration > 2:
                        if event._flag == False:
                            print("任务执行中，长按" + press_time["pause_key"] + "键暂停执行。")
                            print("Task is running, long press '" + press_time["pause_key"] + "' to pause.")
                            # 设置Event的值为True，使得线程b可以继续执行
                            event.set()
                        else:
                            # 设置Event的值为False，使得线程b暂停执行
                            print("任务已暂停，长按" + press_time["pause_key"] + "键继续执行...")
                            print("Task paused, long press '" + press_time["pause_key"] + "' to continue...")
                            event.clear()
                        press_time["duration"] = time.time()
                        press_time["is_pressed"] = False
                    # print("按下p键时间：", press_time["duration"])
        except:
            pass
    return on_press


def on_release_creator(event, press_time):
    def on_release(key):
        try:
            # duration = time.time() - press_time["duration"]
            # # print("松开p键时间：", time.time(), "Duration: ", duration)
            # if duration > 2.5 and key.char == 'p':
            #     if event._flag == False:
            #         print("任务执行中，按p键暂停执行。")
            #         print("Task is running, press 'p' to pause.")
            #         # 设置Event的值为True，使得线程b可以继续执行
            #         event.set()
            #     else:
            #         # 设置Event的值为False，使得线程b暂停执行
            #         print("任务已暂停，按p键继续执行...")
            #         print("Task paused, press 'p' to continue...")
            #         event.clear()
            #     press_time["duration"] = time.time()
            press_time["is_pressed"] = False
        except:
            pass
    return on_release


# def check_pause(key, event):
#     while True:
#         if keyboard.is_pressed(key):  # 按下p键，暂停程序
#             if event._flag == False:
#                 print("任务执行中，长按p键暂停执行。")
#                 print("Task is running, long press 'p' to pause.")
#                 # 设置Event的值为True，使得线程b可以继续执行
#                 event.set()
#             else:
#                 # 设置Event的值为False，使得线程b暂停执行
#                 print("任务已暂停，长按p键继续执行...")
#                 print("Task paused, press 'p' to continue...")
#                 event.clear()
#         time.sleep(1)  # 每秒检查一次

def detect_optimizable(param, ignoreWaitElement=True, waitElement=""):
    try:
        splitLine = param["splitLine"]
    except:
        param["splitLine"] = 0
    if param["beforeJS"] == "" and param["afterJS"] == "" and param["contentType"] <= 1 and param["splitLine"] == 0:
        if param["nodeType"] <= 2:
            if ignoreWaitElement or waitElement == "":
                return True
            else:
                return False
        elif param["nodeType"] == 4: # 如果是图片
            if param["downloadPic"]:
                return False
            else:
                return True
    else:
        return False



def download_image(browser, url, save_directory, element=None):
    # 定义浏览器头信息
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    if url.startswith("data:image"):
        base64_data = url.split(",")[1]
        image_data = base64.b64decode(base64_data)
        # 提取文件名
        file_name = str(uuid.uuid4()) + '.png'
        # 构建保存路径
        save_path = os.path.join(save_directory, file_name)
        # 保存图片到本地
        with open(save_path, 'wb') as file:
            file.write(image_data)
        browser.print_and_log("图片已成功下载到:", save_path)
        browser.print_and_log(
            "The image has been successfully downloaded to:", save_path)
    elif is_valid_url(url):
        try:
            # 提取文件名
            file_name = url.split('/')[-1].split("?")[0]

            # 生成唯一的新文件名
            new_file_name = file_name + '_' + \
                str(uuid.uuid4()) + '_' + file_name

            # 构建保存路径
            save_path = os.path.join(save_directory, new_file_name)
            # 发送 GET 请求获取图片数据，加载浏览器的cookies
            s = requests.session()
            cookies = browser.browser.get_cookies()
            for cookie in cookies:
                s.cookies.set(cookie['name'], cookie['value'])
            response = s.get(url, headers=headers)
            # 检查响应状态码是否为成功状态
            if response.status_code == requests.codes.ok:
                # 保存图片到本地
                with open(save_path, 'wb') as file:
                    file.write(response.content)
                browser.print_and_log("图片已成功下载到:", save_path)
                browser.print_and_log(
                    "The image has been successfully downloaded to:", save_path)
            else:
                # browser.print_and_log(f"直接下载图片失败，状态码为:{response.status_code}，尝试使用Selenium下载图片...")
                # browser.print_and_log(
                    # f"Failed to download image directly, status code is: {response.status_code}, try to download image using Selenium...")
                JS = "var xhr = new XMLHttpRequest(); xhr.open('GET', '" + url +"', true); xhr.responseType = 'blob'; xhr.onload = function() {var reader = new FileReader(); reader.readAsDataURL(xhr.response); reader.onloadend = function() { var base64data = reader.result;}}; xhr.send();"""
                base64data = browser.browser.execute_script(JS)
                if base64data:
                    image_data = base64data.b64decode(base64data.split(",")[1])
                    with open(save_path, 'wb') as file:
                        file.write(image_data)
                    browser.print_and_log("图片已成功下载到:", save_path)
                    browser.print_and_log(
                        "The image has been successfully downloaded to:", save_path)
                else:
                    browser.print_and_log("下载图片失败，只能使用元素截图功能下载图片。")
                    browser.print_and_log("Failed to download image, can only download image using element screenshot function.")
                    # 使用元素截图功能下载图片
                    try:
                        element.screenshot(save_path)
                        browser.print_and_log("图片截图已保存到:", save_path)
                        browser.print_and_log(
                            "The image screenshot has been saved to:", save_path)
                    except Exception as e:
                        browser.print_and_log("下载图片失败|Error downloading image: ", e)
        except Exception as e:
            browser.print_and_log("下载图片失败|Error downloading image: ", e)
    else:
        browser.print_and_log("下载图片失败，请检查此图片链接是否有效:", url)
        browser.print_and_log(
            "Failed to download image, please check if this image link is valid:", url)


def get_output_code(output):
    try:
        if output.find("rue") != -1:  # 如果返回值中包含true
            code = 1
        else:
            code = int(output)
    except:
        code = 0
    return code

# 判断字段是否为空


def isnotnull(s):
    return len(s) != 0


def new_line(outputParameters, maxViewLength, record):
    line = []
    i = 0
    for value in outputParameters.values():
        line.append(value)
        if record[i]:
            print(value[:maxViewLength], " ", end="")
        i += 1
    print("")
    return line


def write_to_csv(file_name, data, record):
    with open(file_name, 'a', encoding='utf-8-sig', newline="") as f:
        f_csv = csv.writer(f)
        for line in data:
            to_write = []
            for i in range(len(line)):
                if record[i]:
                    to_write.append(line[i])
            f_csv.writerow(to_write)
        f.close()

def replace_field_values(orginal_text, outputParameters, browser=None):
    pattern = r'Field\["([^"]+)"\]'
    try:
        replaced_text = re.sub(
            pattern, lambda match: outputParameters.get(match.group(1), ''), orginal_text)
        if re.search(r'eval\(', replaced_text, re.IGNORECASE): # 如果返回值中包含EVAL
            replaced_text = replaced_text.replace("self.", "browser.")
            pattern = re.compile(r'(?i)eval\("(.+?)"\)')
            # 循环替换所有匹配到的eval语句
            while True:
                match = pattern.search(replaced_text)
                if not match:
                    break
                # 执行eval并将其结果转换为字符串形式
                eval_replaced_text = str(eval(match.group(1)))
                # 替换eval语句
                replaced_text = replaced_text.replace(match.group(0), eval_replaced_text)
        if re.search(r'JS\(', replaced_text, re.IGNORECASE): # 如果返回值中包含JS
            replaced_text = replaced_text.replace("self.", "browser.")
            pattern = re.compile(r'(?i)JS\("(.+?)"\)')
            # 循环替换所有匹配到的JS语句
            while True:
                match = pattern.search(replaced_text)
                if not match:
                    break
                # 执行JS并将其结果转换为字符串形式
                JS_replaced_text = str(browser.browser.execute_script(match.group(1)))
                # 替换JS语句
                replaced_text = replaced_text.replace(match.group(0), JS_replaced_text)
    except Exception as e:
        print("eval替换失败，请检查eval语句是否正确。| Failed to replace eval, please check if the eval statement is correct.")
        print(e)
        replaced_text = orginal_text
    return replaced_text


def readCode(code):
    if code.startswith("outside:"):
        file_name = os.path.join(os.path.abspath("./"), code[8:])
        with open(file_name, 'r', encoding='utf-8-sig') as file_obj:
            code = file_obj.read()
    return code

def write_to_json(file_name, data, types, record, keys):
    keys = list(keys)
    # Prepare empty list for data
    data_to_write = []
    # Tranform data and append to list
    for line in data:
        to_write = {}
        for i in range(len(line)):
            if types[i] == "int" or types[i] == "bigInt":
                try:
                    line[i] = int(line[i])
                except:
                    line[i] = 0
            elif types[i] == "double":
                try:
                    line[i] = float(line[i])
                except:
                    line[i] = 0.0
            if record[i]:
                to_write.update({keys[i]: line[i]})
        data_to_write.append(to_write)

    try:
        # read data from JSON
        with open(file_name, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
    except:
        json_data = []

    json_data.extend(data_to_write)

    # write data to JSON
    with open(file_name, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False)


def write_to_excel(file_name, data, types, record):
    # 首先，检查文件是否存在来决定是否处理第一行
    # first = not os.path.exists(file_name)

    # # 准备新数据
    # new_data = pd.DataFrame(data)

    # # 如果不是第一行（即文件已存在），对数据应用类型转换
    # if not first:
    #     for i, col_type in enumerate(types):
    #         if col_type == "int" or col_type == "bigInt":
    #             try:
    #                 new_data[i] = pd.to_numeric(new_data[i], errors='coerce').astype(int)
    #             except:
    #                 new_data[i] = pd.to_numeric("0", errors='coerce').astype(int)
    #         elif col_type == "double":
    #             try:
    #                 new_data[i] = pd.to_numeric(new_data[i], errors='coerce')(0.0)
    #             except:
    #                 new_data[i] = pd.to_numeric("0.0", errors='coerce').astype(float)
    # # 根据 record 筛选列
    # new_data = new_data.loc[:, record]

    # # 如果文件存在，则读取现有数据并追加新数据
    # if first:
    #     combined_data = new_data
    # else:
    #     # 使用 Pandas 读取现有数据
    #     existing_data = pd.read_excel(file_name)
    #     # 合并现有数据与新数据
    #     combined_data = pd.concat([existing_data, new_data], ignore_index=True)

    # # 将合并后的数据写入 Excel
    # combined_data.to_excel(file_name, index=False, engine='openpyxl')

    # existing_data = []
    # first = True
    # # 检查文件是否存在
    # if os.path.exists(file_name):
    #     # 使用 openpyxl 读取现有数据
    #     workbook = load_workbook(file_name, read_only=True)
    #     sheet = workbook.active
    #     # 读取已有行数
    #     num_rows = sheet.max_row
    #     if num_rows > 5000:
    #         print("Excel文件中的数据行数超过5000行，过多的行数将会导致追加模式写入数据速度变慢，建议更换为CSV文件或MySQL数据库存储数据。正在读取数据，请稍等...")
    #         print("The number of rows in the Excel file exceeds 5000, too many rows will cause the speed of writing data in append mode to slow down, it is recommended to replace it with CSV file or MySQL database to store data. Reading data, please wait...")
    #     # existing_data = [[sheet.cell(row=i, column=j).value for j in range(1, sheet.max_column + 1)] for i in range(1, sheet.max_row + 1)]
    #     for i in range(1, sheet.max_row + 1):
    #         row_data = []
    #         if num_rows > 5000 and i % 500 == 0:
    #             print(f"正在读取第{i}/{num_rows}行的数据...")
    #             print(f"Reading data of row {i}/{num_rows}...")
    #         for j in range(1, sheet.max_column + 1):
    #             cell = sheet.cell(row=i, column=j).value
    #             if cell is None:
    #                 cell = ""
    #             row_data.append(cell)
    #         existing_data.append(row_data)
    #     first = False  # 如果文件存在，首行不再是标题行

    # # 使用 xlsxwriter 创建新文件
    # workbook = xlsxwriter.Workbook(file_name)
    # worksheet = workbook.add_worksheet()

    # # 写入现有数据
    # for row_num, row_data in enumerate(existing_data):
    #     for col_num, cell in enumerate(row_data):
    #         worksheet.write(row_num, col_num, cell)

    # # 写入新数据
    # row = len(existing_data)
    # for line in data:
    #     to_write = []
    #     for i in range(len(line)):
    #         value = line[i]
    #         if not first:  # 如果不是第一行，需要转换数据类型
    #             if types[i] == "int" or types[i] == "bigInt":
    #                 try:
    #                     value = int(value)
    #                 except ValueError:
    #                     value = 0
    #             elif types[i] == "double":
    #                 try:
    #                     value = float(value)
    #                 except ValueError:
    #                     value = 0.0
    #         if record[i]:
    #             to_write.append(value)
    #     first = False  # 更新 first 以跳过数据类型转换
    #     for col, item in enumerate(to_write):
    #         worksheet.write(row, col, item)
    #     row += 1

    # # 关闭工作簿
    # workbook.close()

    first = False
    if os.path.exists(file_name):
        # 加载现有的工作簿
        wb = load_workbook(file_name)
        # 行数读取
        num_rows = wb.active.max_row
        if num_rows > 1000:
            print("Excel文件中的数据行数已超过1000行，过多的行数将会导致追加模式写入数据速度变慢，建议增大任务保存对话框中的“每采集多少条数据保存一次”选项的值以提升采集速度，或者更换为CSV文件或MySQL数据库存储数据。正在读取数据，请稍等...")
            print("The number of rows in the Excel file already exceeds 1000, too many rows will cause the speed of writing data in append mode to slow down, it is recommended to increase the value of the 'Save every how many data' option in the task save dialog to improve the collection speed, or replace it with CSV file or MySQL database to store data. Reading data, please wait...")
        ws = wb.active
        if num_rows > 1000:
            print("读取数据完成，正在追加数据...")
            print("Reading data completed, appending data...")
    else:
        # 创建新的工作簿和工作表
        wb = Workbook()
        ws = wb.active
        first = True
    # 追加数据到工作表
    for line in data:
        if not first:  # 如果不是第一行，需要转换数据类型
            for i in range(len(line)):
                if types[i] == "int" or types[i] == "bigInt":
                    try:
                        line[i] = int(line[i])
                    except:
                        line[i] = 0
                elif types[i] == "double":
                    try:
                        line[i] = float(line[i])
                    except:
                        line[i] = 0.0
        else:
            first = False
        to_write = []
        for i in range(len(line)):
            if record[i]:
                to_write.append(line[i])
        try:
            ws.append(to_write)
        except:
            print("写入Excel文件失败，请检查数据类型是否正确。")
            print("Failed to write to Excel file, please check if the data type is correct.")
    # 保存工作簿
    try:
        wb.save(file_name)
    except:
        print("保存Excel文件失败，请检查文件是否被其他程序打开。")
        print("Failed to save Excel file, please check if the file is opened by other programs.")

class Time:
    def __init__(self, type1=""):
        self.t = int(round(time.time() * 1000))
        self.type = type1

    def end(self):
        at = int(round(time.time() * 1000))
        print("Time used for", self.type, ":", at - self.t, "ms")


class myMySQL:
    def __init__(self, config_file="mysql_config.json"):
        # 读取配置文件
        try:
            if sys.platform == "darwin":
                if config_file.find("./") >= 0:
                    config_file = config_file.replace("./", "")
                config_file = os.path.expanduser(
                    "~/Library/Application Support/EasySpider/" + config_file)
            print("MySQL config file path: ", config_file)
            with open(config_file, 'r') as f:
                config = json.load(f)
                self.host = config["host"]
                self.port = config["port"]
                self.username = config["username"]
                self.password = config["password"]
                self.db = config["database"]
        except Exception as e:
            print("读取配置文件失败，请检查配置文件："+config_file+"是否存在，或配置信息是否有误。")
            print("Failed to read configuration file, please check if the configuration file: " +
                  config_file+" exists, or if the configuration information is incorrect.")
            print(e)
        self.connect()
        
    def connect(self):
        try:
            self.conn = pymysql.connect(
                host=self.host, port=self.port, user=self.username, passwd=self.password, db=self.db)
            print("成功连接到数据库。")
            print("Successfully connected to the database.")
        except:
            print("连接数据库失败，请检查配置文件是否正确。")
            print(
                "Failed to connect to the database, please check if the configuration file is correct.")
            sys.exit()

    def create_table(self, table_name, parameters, remove_if_exists=False):
        self.table_name = table_name
        self.field_sql = "("
        self.cursor = self.conn.cursor()
        # 检查表是否存在
        self.cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
        result = self.cursor.fetchone()
        # 如果表存在，删除它
        if result and remove_if_exists:
            self.cursor.execute(f"DROP TABLE {table_name}")
            result = None
            print(f'数据表 {table_name} 已存在，已删除。')

        sql = "CREATE TABLE " + table_name + \
            " (_id INT AUTO_INCREMENT PRIMARY KEY, "
        for item in parameters:
            try:
                recordASField = item["recordASField"]
            except:
                item["recordASField"] = True
            if item["recordASField"]:
                name = item['name']
                if item['type'] == 'int':
                    sql += f"{name} INT, "
                elif item['type'] == 'double':
                    sql += f"{name} DOUBLE, "
                elif item['type'] == 'text':
                    sql += f"{name} TEXT, "
                elif item['type'] == 'mediumText':
                    sql += f"{name} MEDIUMTEXT, "
                elif item['type'] == 'longText':
                    sql += f"{name} LONGTEXT, "
                elif item['type'] == 'datetime':
                    sql += f"{name} DATETIME, "
                elif item['type'] == 'date':
                    sql += f"{name} DATE, "
                elif item['type'] == 'time':
                    sql += f"{name} TIME, "
                elif item['type'] == 'varchar':
                    sql += f"{name} VARCHAR(255), "
                elif item['type'] == 'bigInt':
                    sql += f"{name} BIGINT, "
                self.field_sql += f"{name}, "
        # 移除最后的逗号并添加闭合的括号
        sql = sql.rstrip(', ') + ")"
        self.field_sql = self.field_sql.rstrip(', ') + ")"

        # 如果表不存在，创建它
        if not result:
            # 执行SQL命令
            self.cursor.execute(sql)
        else:
            print(f'数据表 {table_name} 已存在')
            print(f'The data table {table_name} already exists.')
        self.cursor.close()

    def write_to_mysql(self, OUTPUT, record, types):
        # 创建一个游标对象
        self.cursor = self.conn.cursor()

        for line in OUTPUT:
            for i in range(len(line)):
                if types[i] == "int" or types[i] == "bigInt":
                    try:
                        line[i] = int(line[i])
                    except Exception as e:
                        print(e)
                        line[i] = 0
                elif types[i] == "double":
                    try:
                        line[i] = float(line[i])
                    except Exception as e:
                        print(e)
                        line[i] = 0.0
                elif types[i] == "datetime":
                    try:
                        line[i] = datetime.datetime.strptime(
                            line[i], '%Y-%m-%d %H:%M:%S')
                    except Exception as e:
                        print(e)
                        line[i] = datetime.datetime.strptime(
                            "1970-01-01 00:00:00", '%Y-%m-%d %H:%M:%S')
                elif types[i] == "date":
                    try:
                        line[i] = datetime.datetime.strptime(
                            line[i], '%Y-%m-%d')
                    except Exception as e:
                        print(e)
                        line[i] = datetime.datetime.strptime(
                            "1970-01-01", '%Y-%m-%d')
                elif types[i] == "time":
                    try:
                        line[i] = datetime.datetime.strptime(
                            line[i], '%H:%M:%S')
                    except Exception as e:
                        print(e)
                        line[i] = datetime.datetime.strptime(
                            "00:00:00", '%H:%M:%S')
            to_write = []
            for i in range(len(line)):
                if record[i]:
                    to_write.append(line[i])
            # 构造插入数据的 SQL 语句, IGNORE表示如果主键重复则忽略
            sql = f'INSERT IGNORE INTO {self.table_name} {self.field_sql} VALUES ('
            for _ in to_write:
                sql += "%s, "
            # 移除最后的逗号并添加闭合的括号
            sql = sql.rstrip(', ') + ")"
            # 执行 SQL 语句
            try:
                self.cursor.execute(sql, to_write)
            except pymysql.OperationalError as e:
                print("Error:", e)
                print("Try to reconnect to the database...")
                self.connect()
                self.cursor = self.conn.cursor() # 重新创建游标对象
                self.cursor.execute(sql, to_write) # 重新执行SQL语句
                # self.write_to_mysql(OUTPUT, record, types)
            except Exception as e:
                print("Error:", e)
                print("Error SQL:", sql, to_write)
                print("插入数据库错误，请查看以上的错误提示，然后检查数据的类型是否正确，是否文本过长（超过一万的文本类型要设置为大文本）。")
                print("Inserting database error, please check the above error, and then check whether the data type is correct, whether the text is too long (text type over 10,000 should be set to large text).")
                print("重新执行任务时，请删除数据库中的数据表" + self.table_name + "，然后再次运行程序。")
                print("When re-executing the task, please delete the data table " +
                      self.table_name + " in the database, and then run the program again.")

        # 提交到数据库执行
        self.conn.commit()

        # 关闭游标和连接
        self.cursor.close()

    def remove_duplicate_data(self):
        self.cursor = self.conn.cursor()
        # 删除重复数据
        fields = self.field_sql.replace("(", "").replace(")", "")
        sql = f"CREATE TABLE {self.table_name}_temp AS " + \
        f"SELECT MIN(_id) AS _id, " + fields + \
        f" FROM {self.table_name} GROUP BY " + fields + ";"
        self.cursor.execute(sql)
        sql = f"DELETE FROM {self.table_name};"
        self.cursor.execute(sql)
        sql = f"INSERT INTO {self.table_name} SELECT * FROM {self.table_name}_temp;"
        self.cursor.execute(sql)
        sql = f"DROP TABLE {self.table_name}_temp;"
        self.cursor.execute(sql)
        # 提交到数据库执行
        self.conn.commit()
        # 关闭游标和连接
        self.cursor.close()

    def close(self):
        try:
            self.conn.close()
            print("成功关闭数据库。")
            print("Successfully closed the database.")
        except:
            print("关闭数据库失败。")
            print("Failed to close the database.")
    
    def __del__(self):
        self.close()
