# -*- coding: utf-8 -*-
# import atexit
import atexit
import copy
import platform
import shutil
import string
import threading
# import undetected_chromedriver as uc
from utils import detect_optimizable, download_image, extract_text_from_html, get_output_code, isnotnull, lowercase_tags_in_xpath, myMySQL, new_line, \
    on_press_creator, on_release_creator, readCode, rename_downloaded_file, replace_field_values, send_email, split_text_by_lines, write_to_csv, write_to_excel, write_to_json
from constants import WriteMode, DataWriteMode, GraphOption
from myChrome import MyChrome
from threading import Thread, Event
from PIL import Image
from commandline_config import Config
import os
import csv
from openpyxl import load_workbook, Workbook
import random
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import StaleElementReferenceException, InvalidSelectorException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from datetime import datetime
import io  # 遇到错误退出时应执行的代码
import json
# from lib2to3.pgen2 import driver
import re
# import shutil
import subprocess
import sys
# from urllib import parse
# import base64
# import hashlib
import time
import requests
from http.server import BaseHTTPRequestHandler, HTTPServer
from multiprocessing import freeze_support
freeze_support()  # 防止无限死循环多开
try:
    from ddddocr import DdddOcr
    import onnxruntime
    onnxruntime.set_default_logger_severity(3)  # 隐藏onnxruntime的日志
except:
    print("OCR识别无法在当前环境下使用（ddddocr库缺失），请使用完整版执行器easyspider_executestage_full来运行需要OCR识别的任务。")
    print("OCR recognition cannot be used in the current environment (ddddocr library is missing), please use the executor with ddddocr 'easyspider_executestage_full' to run the task which requires OCR recognition.")
from urllib.parse import urljoin
from lxml import etree, html
try:
    import pandas as pd
except:
    print("数据去重无法在当前环境下使用（pandas库缺失），请使用完整版执行器easyspider_executestage_full来运行需要去重的任务。")
    print("Data deduplication cannot be used in the current environment (pandas library is missing), please use the executor with pandas 'easyspider_executestage_full' to run the task which requires data deduplication.")
    time.sleep(1)

# import numpy
# import pytesseract
# import uuid
if sys.platform != "darwin":
    from myChrome import MyUCChrome
desired_capabilities = DesiredCapabilities.CHROME
desired_capabilities["pageLoadStrategy"] = "none"


class BrowserThread(Thread):
    def __init__(self, browser_t, id, service, version, event, saveName, config, option, shutdown_event, commandline_config=""):
        Thread.__init__(self)
        self.logs = io.StringIO()
        # 退出事件，用于远程执行时的中断
        self.shutdown_event = shutdown_event
        self.log = bool(service.get("recordLog", True))
        self.browser = browser_t
        self.option = option
        self.commandline_config = commandline_config
        self.version = version
        self.totalSteps = 0
        self.id = id
        self.event = event
        now = datetime.now()
        self.saveName = service.get("saveName", now.strftime("%Y_%m_%d_%H_%M_%S"))  # 保存文件的名字
        self.OUTPUT = ""
        self.SAVED = False
        self.BREAK = False
        self.CONTINUE = False
        self.browser.maximize_window() if service.get("maximizeWindow") == 1 else ...
        # 名称设定
        if saveName != "":  # 命令行覆盖保存名称
            self.saveName = saveName  # 保存文件的名字
        now = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
        self.saveName = self.saveName.replace("current_time", now)

        self.print_and_log("任务ID", id, "的保存文件名为:", self.saveName)
        self.print_and_log("Save Name for task ID", id, "is:", self.saveName)
        if not os.path.exists("Data/Task_" + str(id)):
            os.mkdir("Data/Task_" + str(id))
        self.downloadFolder = "Data/Task_" + str(id) + "/" + self.saveName
        if not os.path.exists(self.downloadFolder):
            os.mkdir(self.downloadFolder)  # 创建保存文件夹用来保存截图和文件
        if not os.path.exists(self.downloadFolder + "/files"):
            os.mkdir(self.downloadFolder + "/files")
        if not os.path.exists(self.downloadFolder + "/images"):
            os.mkdir(self.downloadFolder + "/images")
        if not os.path.exists(self.downloadFolder + "/screenshots"):
            os.mkdir(self.downloadFolder + "/screenshots")
        self.getDataStep = 0
        self.startSteps = 0
        try:
            if service.get("startFromExit", 0) == 1:
                with open("Data/Task_" + str(self.id) + "/" + self.saveName + '_steps.txt', 'r',
                          encoding='utf-8-sig') as file_obj:
                    self.startSteps = int(file_obj.read())  # 读取已执行步数
        except Exception as e:
            self.print_and_log(f"读取steps.txt失败，原因：{str(e)}")

        if self.startSteps != 0:
            self.print_and_log("此模式下，任务ID", self.id, "将从上次退出的步骤开始执行，之前已采集条数为",
                               self.startSteps, "条。")
            self.print_and_log("In this mode, task ID", self.id,
                               "will start from the last step, before we already collected", self.startSteps, " items.")
        else:
            self.print_and_log("此模式下，任务ID", self.id,
                               "将从头开始执行，如果需要从上次退出的步骤开始执行，请在保存任务时设置是否从上次保存位置开始执行为“是”。")
            self.print_and_log("In this mode, task ID", self.id,
                               "will start from the beginning, if you want to start from the last step, please set the option 'start from the last step' to 'yes' when saving the task.")
        stealth_path = driver_path[:driver_path.find(
            "chromedriver")] + "stealth.min.js"
        with open(stealth_path, 'r') as f:
            js = f.read()
            self.print_and_log("Loading stealth.min.js")
        self.browser.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': js})  # TMALL 反扒
        self.browser.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """
                Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
                })
        """
        })
        WebDriverWait(self.browser, 10)
        self.browser.command_executor._commands["send_command"] = ("POST", '/session/$sessionId/chromium/send_command')
        path = os.path.join(os.path.abspath("./"), "Data", "Task_" + str(self.id), self.saveName, "files")
        self.paramss = {'cmd': 'Page.setDownloadBehavior', 'params': {'behavior': 'allow', 'downloadPath': path}}
        self.browser.execute("send_command", self.paramss)  # 下载目录改变
        self.monitor_event = threading.Event()
        self.monitor_thread = threading.Thread(target=rename_downloaded_file, args=(path, self.monitor_event)) #path后面的逗号不能省略，是元组固定写法
        self.monitor_thread.start()
        # self.browser.get('about:blank')
        self.procedure = service["graph"]  # 程序执行流程
        self.maxViewLength = service.get("maxViewLength", 15)  # 最大显示长度
        self.outputFormat = service.get("outputFormat", "csv")  # 输出格式
        self.save_threshold = service.get("saveThreshold", 10)  # 保存最低阈值
        self.dataWriteMode = service.get("dataWriteMode", DataWriteMode.Append.value)  # 数据写入模式，1为追加，2为覆盖，3为重命名文件
        self.task_version = service.get("version", "")  # 任务版本

        if not self.task_version:
            self.print_and_log("版本不一致，请使用v0.2.0版本的EasySpider运行该任务！")
            self.print_and_log("Version not match, please use EasySpider v0.2.0 to run this task!")
            self.browser.quit()
            sys.exit()

        if self.task_version >= "0.3.1":  # 0.3.1及以上版本以上的EasySpider兼容从0.3.1版本开始的所有版本
            pass
        elif self.task_version != version:  # 0.3.1以下版本的EasySpider不兼容0.3.1及以上版本的EasySpider
            self.print_and_log(f"版本不一致，请使用{self.task_version}版本的EasySpider运行该任务！")
            self.print_and_log(f"Version not match, please use EasySpider {self.task_version} to run this task!")
            self.browser.quit()
            sys.exit()

        service_links = service.get("links")
        if service_links:
            self.links = list(filter(isnotnull, service_links.split("\n")))  # 要执行的link的列表
        else:
            self.links = list(filter(isnotnull, service["url"]))  # 要执行的link

        self.OUTPUT = []  # 采集的数据
        if self.outputFormat in ["csv", "txt", "xlsx", "json"]:
            if os.path.exists("Data/Task_" + str(self.id) + "/" + self.saveName + '.' + self.outputFormat):
                if self.dataWriteMode == DataWriteMode.Cover.value:
                    os.remove("Data/Task_" + str(self.id) + "/" + self.saveName + '.' + self.outputFormat)
                elif self.dataWriteMode == DataWriteMode.Rename.value:
                    i = 2
                    while os.path.exists("Data/Task_" + str(self.id) + "/" + self.saveName + '_' + str(i) + '.' + self.outputFormat):
                        i = i + 1
                    self.saveName = self.saveName + '_' + str(i)
                    self.print_and_log("文件已存在，已重命名为", self.saveName)
        self.writeMode = WriteMode.Append.value   # 写入模式，0为新建，1为追加
        if self.outputFormat in ['csv', 'txt', 'xlsx']:
            if not os.path.exists(f"Data/Task_{str(self.id)}/{self.saveName}.{self.outputFormat}"):
                self.OUTPUT.append([])  # 添加表头
                self.writeMode = WriteMode.Create.value
        elif self.outputFormat == "json":
            self.writeMode = WriteMode.Json.value  # JSON模式无需判断是否存在文件
        elif self.outputFormat == "mysql":
            self.mysql = myMySQL(config["mysql_config_path"])
            self.mysql.create_table(self.saveName, service["outputParameters"],
                                    remove_if_exists=self.dataWriteMode == DataWriteMode.Cover.value)
            self.writeMode = WriteMode.MySQL.value  # MySQL模式

        if self.writeMode == WriteMode.Create.value:
            self.print_and_log("新建模式|Create Mode")
        elif self.writeMode == WriteMode.Append.value:
            self.print_and_log("追加模式|Append Mode")
        elif self.writeMode == WriteMode.MySQL.value:
            self.print_and_log("MySQL模式|MySQL Mode")
        elif self.writeMode == WriteMode.Json.value:
            self.print_and_log("JSON模式|JSON Mode")

        self.containJudge = service["containJudge"]  # 是否含有判断语句
        self.outputParameters = {}
        self.service = service
        self.outputParametersTypes = []
        self.outputParametersRecord = []  # 字段是否被记录
        self.dataNotFoundKeys = {}  # 记录没有找到数据的key
        self.history = {"index": 0, "handle": None}  # 记录页面现在所以在的历史记录的位置
        self.SAVED = False  # 记录是否已经存储了
        for param in service["outputParameters"]:  # 初始化输出参数
            if param["name"] not in self.outputParameters.keys():
                self.outputParameters[param["name"]] = ""
                self.dataNotFoundKeys[param["name"]] = False
                self.outputParametersTypes.append(param.get("type", "text"))
                self.outputParametersRecord.append(bool(param.get("recordASField", True)))
                # 文件叠加的时候不添加表头
                if self.outputFormat in ["csv", "txt", "xlsx"] and self.writeMode == WriteMode.Create.value:
                    self.OUTPUT[0].append(param["name"])
        self.urlId = 0  # 全局记录变量
        self.preprocess()  # 预处理，优化提取数据流程
        self.inputExcel = service.get("inputExcel", "")  # 输入Excel
        self.readFromExcel()  # 读取Excel获得参数值

    # 检测如果没有复杂的操作，优化提取数据流程
    def preprocess(self):
        for index_node, node in enumerate(self.procedure):
            parameters: dict = node["parameters"]
            iframe = parameters.get('iframe')
            option = node["option"]

            parameters["iframe"] = False if not iframe else parameters.get('iframe', False)
            if parameters.get("xpath"):
                parameters["xpath"] = lowercase_tags_in_xpath(parameters["xpath"])

            if parameters.get("waitElementIframeIndex"):
                parameters["waitElementIframeIndex"] = int(parameters["waitElementIframeIndex"])
            else:
                parameters["waitElement"] = ""
                parameters["waitElementTime"] = 10
                parameters["waitElementIframeIndex"] = 0

            if option == GraphOption.Get.value:  # 打开网页操作
                parameters["cookies"] = parameters.get("cookies", "")
            elif option == GraphOption.Click.value:  # 点击操作
                parameters["alertHandleType"] = parameters.get("alertHandleType", 0)
                if parameters.get("useLoop"):
                    if self.task_version <= "0.3.5":
                        # 0.3.5及以下版本的EasySpider下的循环点击不支持相对XPath
                        parameters["xpath"] = ""
                        self.print_and_log(f"您的任务版本号为{self.task_version}，循环点击不支持相对XPath写法，已自动切换为纯循环的XPath")
            elif option == GraphOption.Extract.value:  # 提取数据操作
                parameters["recordASField"] = 0
                parameters["params"] = parameters.get("params", parameters.get("paras"))  # 兼容0.5.0及以下版本的EasySpider
                parameters["clear"] = parameters.get("clear", 0)
                parameters["newLine"] = parameters.get("newLine", 1)

                params = parameters["params"]
                for param in params:
                    param["iframe"] = param.get("iframe", False)

                    if param.get("relativeXPath"):
                        param["relativeXPath"] = lowercase_tags_in_xpath(param["relativeXPath"])

                    parameters["recordASField"] = param.get("recordASField", 1)

                    param["splitLine"] = 0 if not param.get("splitLine") else param.get("splitLine")

                    if param.get("contentType") == 8:
                        self.print_and_log("默认的ddddocr识别功能如果觉得不好用，可以自行修改源码get_content函数->contentType =="
                                           "8的位置换成自己想要的OCR模型然后自己编译运行；或者可以先设置采集内容类型为“元素截图”把图片"
                                           "保存下来，然后用自定义操作调用自己写的程序，程序的功能是读取这个最新生成的图片，然后用好用"
                                           "的模型，如PaddleOCR把图片识别出来，然后把返回值返回给程序作为参数输出。")
                        self.print_and_log("If you think the default ddddocr function is not good enough, you can "
                                           "modify the source code get_content function -> contentType == 8 position "
                                           "to your own OCR model and then compile and run it; or you can first set "
                                           "the content type of the crawler to \"Element Screenshot\" to save the "
                                           "picture, and then call your own program with custom operations. The "
                                           "function of the program is to read the latest generated picture, then use "
                                           "a good model, such as PaddleOCR to recognize the picture, and then return "
                                           "the return value as a parameter output to the program.")
                    param["optimizable"] = detect_optimizable(param)
            elif option == GraphOption.Input.value:  # 输入文字
                parameters['index'] = parameters.get('index', 0)
            elif option == GraphOption.Custom.value:  # 自定义操作
                parameters['clear'] = parameters.get('clear', 0)
                parameters['newLine'] = parameters.get('newLine', 1)
            elif option == GraphOption.Move.value:  # 移动到元素
                if parameters.get('useLoop'):
                    if self.task_version <= "0.3.5":  # 0.3.5及以下版本的EasySpider下的循环点击不支持相对XPath
                        parameters["xpath"] = ""
                        self.print_and_log(f"您的任务版本号为{self.task_version}，循环点击不支持相对XPath写法，已自动切换为纯循环的XPath")
            elif option == GraphOption.Loop.value:  # 循环操作
                parameters['exitElement'] = "//body" if not parameters.get('exitElement') or parameters.get('exitElement') == "" else parameters.get('exitElement')
                parameters["quickExtractable"] = False  # 是否可以快速提取
                parameters['skipCount'] = parameters.get('skipCount', 0)

                # 如果（不）固定元素列表循环中只有一个提取数据操作，且提取数据操作的提取内容为元素截图，那么可以快速提取
                if len(node["sequence"]) == 1 and self.procedure[node["sequence"][0]]["option"] == 3 \
                        and (int(node["parameters"]["loopType"]) == 1 or int(node["parameters"]["loopType"]) == 2):
                    params = self.procedure[node["sequence"][0]].get("parameters").get("params")
                    if not params:
                        params = self.procedure[node["sequence"][0]]["parameters"]["paras"]  # 兼容0.5.0及以下版本的EasySpider

                    waitElement = self.procedure[node["sequence"][0]]["parameters"].get("waitElement", "")

                    if parameters["iframe"]:
                        parameters["quickExtractable"] = False  # 如果是iframe，那么不可以快速提取
                    else:
                        parameters["quickExtractable"] = True  # 先假设可以快速提取

                    if parameters["skipCount"] > 0:
                        parameters["quickExtractable"] = False  # 如果有跳过的元素，那么不可以快速提取

                    for param in params:
                        optimizable = detect_optimizable(param, ignoreWaitElement=False, waitElement=waitElement)
                        param['iframe'] = param.get('iframe', False)
                        if param["iframe"] and not param["relative"]:  # 如果是iframe，那么不可以快速提取
                            optimizable = False
                        if not optimizable:  # 如果有一个不满足优化条件，那么就不能快速提取
                            parameters["quickExtractable"] = False
                            break

                    if parameters["quickExtractable"]:
                        self.print_and_log(f"循环操作<{node['title']}>可以快速提取数据")
                        self.print_and_log(f"Loop operation <{node['title']}> can extract data quickly")
                        parameters["clear"] = self.procedure[node["sequence"][0]]["parameters"].get("clear", 0)
                        parameters["newLine"] = self.procedure[node["sequence"][0]]["parameters"].get("newLine", 1)

                        if int(node["parameters"]["loopType"]) == 1:  # 不固定元素列表
                            node["parameters"]["baseXPath"] = node["parameters"]["xpath"]
                        elif int(node["parameters"]["loopType"]) == 2:  # 固定元素列表
                            node["parameters"]["baseXPath"] = node["parameters"]["pathList"]
                        node["parameters"]["quickParams"] = []
                        for param in params:
                            content_type = ""
                            if param["relativeXPath"].find("/@href") >= 0 or param["relativeXPath"].find("/text()") >= 0 \
                                    or param["relativeXPath"].find("::text()") >= 0:
                                content_type = ""
                            elif param["nodeType"] == 2:
                                content_type = "//@href"
                            elif param["nodeType"] == 4:  # 图片链接
                                content_type = "//@src"
                            elif param["contentType"] == 1:
                                content_type = "/text()"
                            elif param["contentType"] == 0:
                                content_type = "//text()"
                            if param["relative"]:  # 如果是相对XPath
                                xpath = "." + param["relativeXPath"] + content_type
                            else:
                                xpath = param["relativeXPath"] + content_type
                            # 如果是id()或(//div)[1]这种形式，不需要包/html/body
                            # if xpath.find("/body") < 0 and xpath.startswith("/"):
                            #     xpath = "/html/body" + xpath
                            node["parameters"]["quickParams"].append({
                                "name": param["name"],
                                "relative": param["relative"],
                                "xpath": xpath,
                                "nodeType": param["nodeType"],
                                "default": param["default"],
                            })
            self.procedure[index_node]["parameters"] = parameters
        self.print_and_log("预处理完成|Preprocess completed")

    def readFromExcel(self):
        if self.inputExcel == "":
            return 0
        try:
            workbook = load_workbook(self.inputExcel)
        except:
            self.print_and_log("读取Excel失败，将会使用默认参数执行任务，请检查文件路径是否正确：",
                               os.path.abspath(self.inputExcel))
            self.print_and_log(
                "Failed to read Excel, will execute the task with default parameters, please check if the file path is correct: ",
                os.path.abspath(self.inputExcel))
            time.sleep(5)
            return 0

        sheet_name_list = workbook.sheetnames
        sheet = workbook[sheet_name_list[0]]
        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        result = list(zip(*data))
        result_dict = {}
        for row in result:
            key = row[0]
            values = [str(val) for val in row[1:] if val is not None]
            result_dict.setdefault(key, []).extend([values])

        data = {}
        for key, arr in result_dict.items():
            result = []
            for cols in zip(*arr):
                result.append("~".join(cols))
            data[key] = result

        try:
            if "urlList_0" in data.keys():
                self.links = data["urlList_0"]
        except:
            self.links = "about:blank"
        task = self.service
        for key, value in data.items():
            for i in range(len(task["inputParameters"])):
                if key == task["inputParameters"][i]["name"]:
                    nodeId = int(task["inputParameters"][i]["nodeId"])
                    node = task["graph"][nodeId]
                    value = "\r\n".join(value)
                    if node["option"] == 1:
                        node["parameters"]["links"] = value
                    elif node["option"] == 4:
                        node["parameters"]["value"] = value
                    elif node["option"] == 8 and node["parameters"]["loopType"] == 0:
                        node["parameters"]["exitCount"] = int(value)
                    elif node["option"] == 8:
                        node["parameters"]["textList"] = value
                    break
        self.print_and_log("已从Excel读取输入参数，覆盖了原有输入参数。")
        self.print_and_log(
            "Already read input parameters from Excel and overwrite the original input parameters.")

    def removeDuplicateData(self):
        try:
            removeDuplicateData = self.service["removeDuplicate"]
        except:
            removeDuplicateData = 0
        if removeDuplicateData == 1:
            self.print_and_log("正在去除重复数据，请稍后……")
            self.print_and_log("Removing duplicate data, please wait...")
            if self.outputFormat == "csv" or self.outputFormat == "txt" or self.outputFormat == "json" or self.outputFormat == "xlsx":
                file_name = "Data/Task_" + \
                            str(self.id) + "/" + self.saveName + \
                            '.' + self.outputFormat
                if self.outputFormat == "csv" or self.outputFormat == "txt":
                    df = pd.read_csv(file_name)
                    df.drop_duplicates(inplace=True)
                    df.to_csv(file_name, index=False)
                elif self.outputFormat == "xlsx":
                    df = pd.read_excel(file_name)
                    df.drop_duplicates(inplace=True)
                    df.to_excel(file_name, index=False)
                elif self.outputFormat == "json":
                    df = pd.read_json(file_name)
                    df.drop_duplicates(inplace=True)
                    df.to_json(file_name, orient="records", force_ascii=False)
            elif self.outputFormat == "mysql":
                self.mysql.remove_duplicate_data()
            self.print_and_log("去重完成。")
            self.print_and_log("Duplicate data removed.")

    def run(self):
        # 挨个执行程序
        for i in range(len(self.links)):
            if self.shutdown_event.is_set():
                self.print_and_log("接收到终止信号，正在中断任务... | Received termination signal, interrupting task...")
                break
            self.event.wait()  # 暂停/恢复
            self.executeNode(self.startSteps, self.links[i], "", i)
        # files = os.listdir("Data/Task_" + str(self.id) + "/" + self.saveName)
        # 如果目录为空，则删除该目录
        # if not files:
        #     os.rmdir("Data/Task_" + str(self.id) + "/" + self.saveName)
        if self.shutdown_event.is_set():
            self.print_and_log("任务已中断 | Task interrupted")
        else:
            self.print_and_log("Done!")
            self.print_and_log("执行完成！")
            self.saveData(exit=True)
            self.removeDuplicateData()
            
        if self.outputFormat == "mysql":
            self.mysql.close()
        
        if not self.shutdown_event.is_set():
            try:
                quitWaitTime = self.service["quitWaitTime"]
            except:
                quitWaitTime = 60
            self.print_and_log(f"任务执行完毕，将在{quitWaitTime}秒后自动退出浏览器并清理临时用户目录，等待时间可在保存任务对话框中设置。")
            self.print_and_log(f"The task is completed, the browser will exit automatically and the temporary user directory will be cleaned up after {quitWaitTime} seconds, the waiting time can be set in the save task dialog.")
            
            # 使退出前的等待可被中断
            wait_end_time = time.time() + quitWaitTime
            while time.time() < wait_end_time:
                if self.shutdown_event.is_set():
                    break
                time.sleep(0.1)

        try:
            self.browser.quit()
        except:
            pass
        self.print_and_log("正在清理临时用户目录……|Cleaning up temporary user directory...")
        try:
            shutil.rmtree(self.option["tmp_user_data_folder"])
        except:
            pass
        self.monitor_event.set()
        self.print_and_log("清理完成！|Clean up completed!")
        self.print_and_log("您现在可以安全的关闭此窗口了。|You can safely close this window now.")
        
    def recordLog(self, *args, **kwargs):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
        print(now + ":", *args, file=self.logs, **kwargs)

    # 定义一个自定义的 print 函数，它将内容同时打印到屏幕和文件中
    def print_and_log(self, *args, **kwargs):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
        # 将内容打印到屏幕，立刻输出
        print(*args, **kwargs, flush=True)

        # 将内容写入文件
        print(now + ":", *args, file=self.logs, **kwargs)

    def saveData(self, exit=False):
        # 每save_threshold条保存一次
        if exit == True or len(self.OUTPUT) >= self.save_threshold:
            # 写入日志
            # self.recordLog("持久化存储数据/Persistently store data")
            if self.log:
                with open("Data/Task_" + str(self.id) + "/" + self.saveName + '.log', 'a',
                          encoding='utf-8-sig') as file_obj:
                    file_obj.write(self.logs.getvalue())
                    file_obj.close()
            # 写入已执行步数
            with open("Data/Task_" + str(self.id) + "/" + self.saveName + '_steps.txt', 'w',
                      encoding='utf-8-sig') as file_obj:
                file_obj.write(str(self.totalSteps + 1))
                file_obj.close()
            # 写入数据
            if self.outputFormat == "csv" or self.outputFormat == "txt":
                file_name = "Data/Task_" + \
                            str(self.id) + "/" + self.saveName + \
                            '.' + self.outputFormat
                write_to_csv(file_name, self.OUTPUT,
                             self.outputParametersRecord)
            elif self.outputFormat == "xlsx":
                file_name = "Data/Task_" + \
                            str(self.id) + "/" + self.saveName + '.xlsx'
                write_to_excel(
                    file_name, self.OUTPUT, self.outputParametersTypes, self.outputParametersRecord)
            elif self.outputFormat == "json":
                file_name = "Data/Task_" + \
                            str(self.id) + "/" + self.saveName + '.json'
                write_to_json(file_name, self.OUTPUT, self.outputParametersTypes,
                              self.outputParametersRecord, self.outputParameters.keys())
            elif self.outputFormat == "mysql":
                self.mysql.write_to_mysql(
                    self.OUTPUT, self.outputParametersRecord, self.outputParametersTypes)

            self.OUTPUT = []
            self.logs.truncate(0)  # 清空日志
            self.logs.seek(0)  # 清空日志

    def scrollDown(self, param, rt=""):
        try:
            time.sleep(param["scrollWaitTime"])  # 下拉前等待
        except:
            pass
        scrollType = int(param["scrollType"])
        try:
            param["scrollCount"] = int(param["scrollCount"])
        except:
            param["scrollCount"] = 1
        try:
            if scrollType != 0 and param["scrollCount"] > 0:  # 控制屏幕向下滚动
                if scrollType == 1 or scrollType == 2:
                    for i in range(param["scrollCount"]):
                        body = self.browser.find_element(
                            By.CSS_SELECTOR, "body", iframe=param["iframe"])
                        if scrollType == 1:
                            body.send_keys(Keys.PAGE_DOWN)
                        elif scrollType == 2:
                            body.send_keys(Keys.END)
                        try:
                            time.sleep(param["scrollWaitTime"])  # 下拉完等待
                        except:
                            pass
                        self.print_and_log("向下滚动，第", i + 1, "次。")
                        self.print_and_log(
                            "Scroll down, the", i + 1, "time.")
                elif scrollType == 3:
                    bodyText = ""
                    i = 0
                    while True:
                        newBodyText = self.browser.find_element(
                            By.CSS_SELECTOR, "body", iframe=False).text
                        if param["iframe"]:  # 如果标记了iframe
                            iframes = self.browser.find_elements(
                                By.CSS_SELECTOR, "iframe", iframe=False)
                            for iframe in iframes:
                                self.browser.switch_to.default_content()
                                self.browser.switch_to.frame(iframe)
                                iframe_text = super(self.browser.__class__, self.browser).find_element(
                                    By.CSS_SELECTOR, "body").text  # 用super调用父类的方法
                                newBodyText += iframe_text
                                self.browser.switch_to.default_content()
                        if newBodyText == bodyText:
                            self.print_and_log("页面已检测不到新内容，停止滚动。")
                            self.print_and_log(
                                "No new content detected on the page, stop scrolling.")
                            break
                        else:
                            bodyText = newBodyText
                        body = self.browser.find_element(
                            By.CSS_SELECTOR, "body", iframe=param["iframe"])
                        body.send_keys(Keys.END)
                        self.print_and_log("滚动到底部，第", i + 1, "次。")
                        self.print_and_log(
                            "Scroll to the bottom, the", i + 1, "time.")
                        i = i + 1
                        try:
                            time.sleep(param["scrollWaitTime"])  # 下拉完等待
                        except:
                            pass
        except Exception as e:
            self.print_and_log("滚动屏幕时出错|Error scrolling screen:", e)
            try:
                self.browser.execute_script('window.stop()')
            except:
                pass
            if scrollType != 0 and param["scrollCount"] > 0:  # 控制屏幕向下滚动
                if scrollType == 1 or scrollType == 2:
                    for i in range(param["scrollCount"]):
                        body = self.browser.find_element(
                            By.CSS_SELECTOR, "body", iframe=param["iframe"])
                        if scrollType == 1:
                            body.send_keys(Keys.PAGE_DOWN)
                        elif scrollType == 2:
                            body.send_keys(Keys.END)
                        try:
                            time.sleep(param["scrollWaitTime"])  # 下拉完等待
                        except:
                            pass
                        self.print_and_log("向下滚动，第", i + 1, "次。")
                        self.print_and_log(
                            "Scroll down, the", i + 1, "time.")
                elif scrollType == 3:
                    bodyText = ""
                    i = 0
                    while True:
                        newBodyText = self.browser.find_element(
                            By.CSS_SELECTOR, "body", iframe=False).text
                        if param["iframe"]:  # 如果标记了iframe
                            iframes = self.browser.find_elements(
                                By.CSS_SELECTOR, "iframe", iframe=False)
                            for iframe in iframes:
                                self.browser.switch_to.default_content()
                                self.browser.switch_to.frame(iframe)
                                iframe_text = super(self.browser.__class__, self.browser).find_element(
                                    By.CSS_SELECTOR, "body").text  # 用super调用父类的方法
                                newBodyText += iframe_text
                                self.browser.switch_to.default_content()
                        if newBodyText == bodyText:
                            self.print_and_log("页面已检测不到新内容，停止滚动。")
                            self.print_and_log(
                                "No new content detected on the page, stop scrolling.")
                            break
                        else:
                            bodyText = newBodyText
                        body = self.browser.find_element(
                            By.CSS_SELECTOR, "body", iframe=param["iframe"])
                        body.send_keys(Keys.END)
                        self.print_and_log("滚动到底部，第", i + 1, "次。")
                        self.print_and_log(
                            "Scroll to the bottom, the", i + 1, "time.")
                        i = i + 1
                        try:
                            time.sleep(param["scrollWaitTime"])  # 下拉完等待
                        except:
                            pass
            if rt != "":
                rt.end()

    def execute_code(self, codeMode, code, max_wait_time, element=None, iframe=False):
        output = ""
        if code == "":
            return ""
        if max_wait_time == 0:
            max_wait_time = 999999
        # self.print_and_log(codeMode, code)
        # 将value中的Field[""]替换为outputParameters中的键值
        code = replace_field_values(code, self.outputParameters, self)
        if iframe and self.browser.iframe_env == False:
            # 获取所有的 iframe
            self.browser.switch_to.default_content()
            iframes = self.browser.find_elements(
                By.CSS_SELECTOR, "iframe", iframe=False)
            # 遍历所有的 iframe 并点击里面的元素
            for iframe in iframes:
                # 切换到 iframe
                try:
                    self.browser.switch_to.default_content()
                    self.browser.switch_to.frame(iframe)
                    self.browser.iframe_env = True
                    break
                except:
                    self.print_and_log("Iframe switch failed")
        elif not iframe and self.browser.iframe_env == True:
            self.browser.switch_to.default_content()
            self.browser.iframe_env = False
        if int(codeMode) == 0:
            self.recordLog("Execute JavaScript:" + code)
            self.recordLog("执行JavaScript:" + code)
            self.browser.set_script_timeout(max_wait_time)
            try:
                output = self.browser.execute_script(code)
            except Exception as e:
                output = ""
                self.print_and_log("执行下面的代码时出错:" + code, "，错误为：", str(e))
                self.print_and_log("Error executing the following code:" + code, ", error is:", str(e))
        elif int(codeMode) == 2:
            self.recordLog("Execute JavaScript for element:" + code)
            self.recordLog("对元素执行JavaScript:" + code)
            self.browser.set_script_timeout(max_wait_time)
            try:
                output = self.browser.execute_script(code, element)
            except Exception as e:
                output = ""
                self.print_and_log("执行下面的代码时出错:" + code, "，错误为：", str(e))
                self.print_and_log("Error executing the following code:" + code, ", error is:", str(e))
        elif int(codeMode) == 5:
            try:
                code = readCode(code)
                # global_namespace = globals().copy()
                # global_namespace["self"] = self
                output = exec(code)
                self.recordLog("执行下面的代码:" + code)
                self.recordLog("Execute the following code:" + code)
            except Exception as e:
                self.print_and_log("执行下面的代码时出错:" + code, "，错误为：", str(e))
                self.print_and_log("Error executing the following code:" +
                                   code, ", error is:", str(e))
        elif int(codeMode) == 6:
            try:
                code = readCode(code)
                output = eval(code)
                self.recordLog("获得下面的代码返回值:" + code)
                self.recordLog(
                    "Get the return value of the following code:" + code)
            except Exception as e:
                self.print_and_log("获得下面的代码返回值时出错:" + code, "，错误为：", e)
                self.print_and_log(
                    "Error executing and getting return value the following code:" + code, ", error is:", e)
        elif int(codeMode) == 1:
            self.recordLog("Execute System Call:" + code)
            self.recordLog("执行系统命令:" + code)
            # 执行系统命令
            try:
                # output = subprocess.run(code, capture_output=True, text=True, timeout=max_wait_time, encoding="utf-8", shell=True)
                output = subprocess.run(
                    code, capture_output=True, text=True, timeout=max_wait_time, shell=True)
                # 输出命令返回值
                output = output.stdout
                self.print_and_log(output)
            except subprocess.TimeoutExpired:
                # 命令执行时间超过指定值，抛出异常
                self.recordLog("Command timed out")
                self.recordLog("命令执行超时")
            except Exception as e:
                self.print_and_log(e)  # 打印异常信息
                self.recordLog("Command execution failed")
                self.recordLog("命令执行失败")
        try:
            output = str(output)
        except:
            output = "无法转换为字符串|Unable to convert to string"
            self.print_and_log("无法转换为字符串|Unable to convert to string", output)
        return output

    def customOperation(self, node, loopValue, loopPath, index):
        params = node["parameters"]
        if params["clear"] == 1:
            self.clearOutputParameters()
        codeMode = int(params["codeMode"])
        code = params["code"]
        output = ""
        max_wait_time = int(params["waitTime"])
        if codeMode == 2:  # 使用循环的情况下，传入的clickPath就是实际的xpath
            try:
                loopPath = replace_field_values(
                    loopPath, self.outputParameters, self)
                elements = self.browser.find_elements(
                    By.XPATH, loopPath, iframe=params["iframe"])
                element = elements[index]
                output = self.execute_code(
                    codeMode, code, max_wait_time, element, iframe=params["iframe"])
            except:
                output = ""
                self.print_and_log("JavaScript execution failed")
        elif codeMode == 3:
            self.BREAK = True
            self.recordLog("跳出循环|Break the loop")
        elif codeMode == 4:
            self.CONTINUE = True
            self.recordLog("跳过本次循环|Skip this loop")
        elif codeMode == 7:  # 暂停程序执行
            self.event.clear()
            self.print_and_log(
                f"根据设置的自定义操作，任务已暂停，长按{self.service['pauseKey']}键继续执行...|Task paused according to custom operation, long press '{self.service['pauseKey']}' to continue...")
        elif codeMode == 8:  # 刷新页面
            self.browser.refresh()
            self.print_and_log("根据设置的自定义操作，任务已刷新页面|Task refreshed page according to custom operation")
        elif codeMode == 9:  # 发送邮件
            send_email(node["parameters"]["emailConfig"])
        elif codeMode == 10: # 清空所有字段值
            self.clearOutputParameters()
        elif codeMode == 11: # 生成新的数据行
            line = new_line(self.outputParameters,
                            self.maxViewLength, self.outputParametersRecord)
            self.OUTPUT.append(line)
        elif codeMode == 12: # 退出程序
            self.print_and_log("根据设置的自定义操作，任务已退出|Task exited according to custom operation")
            self.saveData(exit=True)
            self.browser.quit()
            self.print_and_log("正在清理临时用户目录……|Cleaning up temporary user directory...")
            try:
                shutil.rmtree(self.option["tmp_user_data_folder"])
            except:
                pass
            self.print_and_log("清理完成！|Clean up completed!")
            os._exit(0)
        else:  # 0 1 5 6
            output = self.execute_code(
                codeMode, code, max_wait_time, iframe=params["iframe"])
        recordASField = bool(params["recordASField"])
        # if recordASField:
        # self.print_and_log("操作<" + node["title"] + ">的返回值为：" + output)
        # self.print_and_log("The return value of operation <" + node["title"] + "> is: " + output)
        self.outputParameters[node["title"]] = output
        if recordASField and params["newLine"]:
            line = new_line(self.outputParameters,
                            self.maxViewLength, self.outputParametersRecord)
            self.OUTPUT.append(line)

    def switchSelect(self, param, loopValue):
        optionMode = param["optionMode"]
        optionValue = param["optionValue"]
        if param["useLoop"]:
            index = param["index"]
            if index != 0:
                try:
                    optionValue = loopValue.split("~")[index - 1]
                except:
                    self.print_and_log("取值失败，可能是因为取值索引超出范围，将使用整个文本值")
                    self.print_and_log(
                        "Failed to get value, maybe because the index is out of range, will use the entire text value")
            else:
                optionValue = loopValue
            optionMode = 1
        try:
            xpath = replace_field_values(
                param["xpath"], self.outputParameters, self)
            dropdown = Select(self.browser.find_element(
                By.XPATH, xpath, iframe=param["iframe"]))
            try:
                if optionMode == 0:
                    # 获取当前选中的选项索引
                    current_index = dropdown.options.index(
                        dropdown.first_selected_option)
                    # 计算下一个选项的索引
                    next_index = (current_index + 1) % len(dropdown.options)
                    # 选择下一个选项
                    dropdown.select_by_index(next_index)
                elif optionMode == 1:
                    dropdown.select_by_index(int(optionValue))
                elif optionMode == 2:
                    dropdown.select_by_value(optionValue)
                elif optionMode == 3:
                    dropdown.select_by_visible_text(optionValue)
                # self.recordLog("切换到下拉框选项|Change to drop-down box option:", xpath)
            except:
                self.print_and_log("切换下拉框选项失败:", xpath,
                                   param["optionMode"], param["optionValue"])
                self.print_and_log("Failed to change drop-down box option:",
                                   xpath, param["optionMode"], param["optionValue"])
        except:
            self.print_and_log("找不到下拉框元素:", xpath)
            self.print_and_log("Cannot find drop-down box element:", xpath)

    def moveToElement(self, param, loopElement=None, loopPath="", index=0):
        time.sleep(0.1)  # 移动之前等待0.1秒
        loopPath = replace_field_values(loopPath, self.outputParameters, self)
        xpath = replace_field_values(
            param["xpath"], self.outputParameters, self)
        if param["useLoop"]:  # 使用循环的情况下，传入的clickPath就是实际的xpath
            if xpath == "":
                path = loopPath
            else:
                path = "(" + loopPath + ")" + \
                       "[" + str(index + 1) + "]" + \
                       xpath
                index = 0  # 如果是相对循环内元素的点击，在定位到元素后，index应该重置为0
            # element = loopElement
        else:
            index = 0
            path = xpath  # 不然使用元素定义的xpath
        path = replace_field_values(path, self.outputParameters, self)
        try:
            elements = self.browser.find_elements(
                By.XPATH, path, iframe=param["iframe"])
            element = elements[index]
            try:
                ActionChains(self.browser).move_to_element(element).perform()
                # self.recordLog("移动到元素|Move to element:", path)
            except:
                self.print_and_log("移动鼠标到元素失败:", xpath)
                self.print_and_log("Failed to move mouse to element:", xpath)
        except:
            self.print_and_log("找不到元素:", xpath)
            self.print_and_log("Cannot find element:", xpath)

    # 执行节点关键函数部分
    def executeNode(self, nodeId, loopValue="", loopPath="", index=0):
        if self.shutdown_event.is_set():
            return
        self.event.wait()
        node = self.procedure[nodeId]
        # WebDriverWait(self.browser, 10).until
        # # 等待元素出现才进行操作，10秒内未出现则报错
        # (EC.visibility_of_element_located(
        #     (By.XPATH, node["parameters"]["xpath"])))
        try:
            if node["parameters"]["waitElement"] != "":
                waitElement = replace_field_values(
                    node["parameters"]["waitElement"], self.outputParameters, self)
                waitElementTime = float(node["parameters"]["waitElementTime"])
                waitElementIframeIndex = node["parameters"]["waitElementIframeIndex"]
                self.print_and_log("等待元素出现:", waitElement)
                self.print_and_log(
                    "Waiting for element to appear:", waitElement)
                if waitElementIframeIndex > 0:
                    iframes = self.browser.find_elements(
                        By.CSS_SELECTOR, "iframe", iframe=False)
                    iframe = iframes[waitElementIframeIndex - 1]
                    self.browser.switch_to.frame(iframe)
                WebDriverWait(self.browser, waitElementTime).until(
                    EC.presence_of_element_located((By.XPATH, waitElement))
                )
                if waitElementIframeIndex > 0:
                    self.browser.switch_to.default_content()
        except Exception as e:
            if waitElement != "":
                self.print_and_log("等待元素出现超时：", waitElement, "，将继续执行。")
                self.print_and_log("Timeout waiting for element to appear:",
                                   waitElement, ", will continue to execute.")
                self.recordLog(e)
            self.recordLog("Wait element not found")
        self.recordLog("执行节点|Execute node:", node["title"])
        try:
            # 根据不同选项执行不同操作
            if node["option"] == 0 or node["option"] == 10:  # root操作,条件分支操作
                for i in node["sequence"]:  # 从根节点开始向下读取
                    self.executeNode(i, loopValue, loopPath, index)
            elif node["option"] == 1:  # 打开网页操作
                # if not (nodeId == 1 and self.service["cloudflare"] == 1):
                self.openPage(node["parameters"], loopValue)
            elif node["option"] == 2:  # 点击元素
                self.clickElement(node["parameters"], loopValue, loopPath, index)
            elif node["option"] == 3:  # 提取数据
                # 针对提取数据操作，设置操作开始的步骤，用于不小心关闭后的恢复的增量采集
                if self.totalSteps >= self.startSteps:
                    self.getData(node["parameters"], loopValue, node["isInLoop"],
                                parentPath=loopPath, index=index)
                    self.saveData()
                else:
                    # self.getDataStep += 1
                    self.print_and_log("跳过第" + str(self.totalSteps) + "次提取数据。")
                    self.print_and_log(
                        "Skip the " + str(self.totalSteps) + "th data extraction.")
                self.totalSteps += 1  # 总步数加一
            elif node["option"] == 4:  # 输入文字
                self.inputInfo(node["parameters"], loopValue)
            elif node["option"] == 5:  # 自定义操作
                self.customOperation(node, loopValue, loopPath, index)
                self.saveData()
            elif node["option"] == 6:  # 切换下拉框
                self.switchSelect(node["parameters"], loopValue)
            elif node["option"] == 7:  # 鼠标移动到元素上
                self.moveToElement(node["parameters"], loopValue, loopPath, index)
            elif node["option"] == 8:  # 循环
                self.loopExecute(node, loopValue, loopPath, index)  # 执行循环
            elif node["option"] == 9:  # 条件分支
                self.judgeExecute(node, loopValue, loopPath, index)
        except Exception as e:
            if self.shutdown_event.is_set(): return
            self.print_and_log("执行节点<" + node["title"] + ">时出错，将继续执行，错误为：", e)
            self.print_and_log("Error executing node <" + node["title"] + ">, will continue to execute, error is:", e)
        

        # 执行完之后进行等待
        if node["option"] != 0 and node["option"] != 2:  # 点击元素操作单独定义等待时间操作
            waitTime = 0.01  # 默认等待0.01秒
            if node["parameters"]["wait"] >= 0:
                waitTime = node["parameters"]["wait"]
            try:
                waitType = int(node["parameters"]["waitType"])
            except:
                waitType = 0
            if waitType == 0:  # 固定等待时间
                time.sleep(waitTime)
            elif waitType == 1:  # 随机等待时间
                time.sleep(random.uniform(waitTime * 0.5, waitTime * 1.5))
        self.event.wait()  # 等待事件结束

    # 对判断条件的处理
    def judgeExecute(self, node, loopElement, clickPath="", index=0):
        if self.shutdown_event.is_set():
            return
        self.event.wait()
        executeBranchId = 0  # 要执行的BranchId
        for i in node["sequence"]:
            cnode = self.procedure[i]  # 获得条件分支
            tType = int(cnode["parameters"]["class"])  # 获得判断条件类型
            if tType == 0:  # 什么条件都没有
                executeBranchId = i
                break
            elif tType == 1:  # 当前页面包含文本
                try:
                    bodyText = self.browser.find_element(
                        By.CSS_SELECTOR, "body", iframe=cnode["parameters"]["iframe"]).text
                    value = replace_field_values(
                        cnode["parameters"]["value"], self.outputParameters, self)
                    if bodyText.find(value) >= 0:
                        executeBranchId = i
                        break
                except:  # 找不到元素下一个条件
                    continue
            elif tType == 2:  # 当前页面包含元素
                try:
                    xpath = replace_field_values(
                        cnode["parameters"]["value"], self.outputParameters, self)
                    if self.browser.find_element(By.XPATH, xpath, iframe=cnode["parameters"]["iframe"]):
                        executeBranchId = i
                        break
                except:  # 找不到元素或者xpath写错了，下一个条件
                    continue
            elif tType == 3:  # 当前循环元素包括文本
                try:
                    value = replace_field_values(
                        cnode["parameters"]["value"], self.outputParameters, self)
                    if loopElement.text.find(value) >= 0:
                        executeBranchId = i
                        break
                except:  # 找不到元素或者xpath写错了，下一个条件
                    continue
            elif tType == 4:  # 当前循环元素包括元素
                try:
                    xpath = replace_field_values(
                        cnode["parameters"]["value"][1:], self.outputParameters, self)
                    if loopElement.find_element(By.XPATH, xpath):
                        executeBranchId = i
                        break
                except:  # 找不到元素或者xpath写错了，下一个条件
                    continue
            elif tType <= 8:  # JS命令返回值
                if tType == 5:  # JS命令返回值等于
                    output = self.execute_code(
                        0, cnode["parameters"]["code"], cnode["parameters"]["waitTime"],
                        iframe=cnode["parameters"]["iframe"])
                elif tType == 6:  # System
                    output = self.execute_code(
                        1, cnode["parameters"]["code"], cnode["parameters"]["waitTime"],
                        iframe=cnode["parameters"]["iframe"])
                elif tType == 7:  # 针对当前循环项的JS命令返回值
                    output = self.execute_code(
                        2, cnode["parameters"]["code"], cnode["parameters"]["waitTime"], loopElement,
                        iframe=cnode["parameters"]["iframe"])
                elif tType == 8:  # 针对当前循环项的System命令返回值
                    output = self.execute_code(
                        6, cnode["parameters"]["code"], cnode["parameters"]["waitTime"], loopElement,
                        iframe=cnode["parameters"]["iframe"])
                try:
                    if output.find("rue") != -1:  # 如果返回值中包含true
                        code = 1
                    else:
                        code = int(output)
                except:
                    code = 0
                if code > 0:
                    executeBranchId = i
                    break
        if executeBranchId != 0:
            self.executeNode(executeBranchId, loopElement, clickPath, index)
        else:
            self.recordLog(
                "判断条件内所有条件分支的条件都不满足|None of the conditions in the judgment condition are met")

    def handleHistory(self, node, xpath, thisHandle, thisHistoryURL, thisHistoryLength, index, element=None, elements=None):
        if self.shutdown_event.is_set():
            return
        self.event.wait()
        try:
            changed_handle = self.browser.current_window_handle != thisHandle
        except:  # 如果网页被意外关闭了的情况下
            self.browser.switch_to.window(
                self.browser.window_handles[-1])
            changed_handle = self.browser.window_handles[-1] != thisHandle
        if changed_handle:  # 如果执行完一次循环之后标签页的位置发生了变化
            try:
                while True:  # 一直关闭窗口直到当前标签页
                    self.browser.close()  # 关闭使用完的标签页
                    self.browser.switch_to.window(
                        self.browser.window_handles[-1])
                    if self.browser.current_window_handle == thisHandle:
                        break
            except Exception as e:
                self.print_and_log("关闭标签页发生错误：", e)
                self.print_and_log(
                    "Error occurred while closing tab: ", e)
        if self.history["index"] != thisHistoryLength and self.history["handle"] == self.browser.current_window_handle:  # 如果执行完一次循环之后历史记录发生了变化，注意当前页面的判断
            difference = thisHistoryLength - self.history["index"]  # 计算历史记录变化差值
            self.browser.execute_script('history.go(' + str(difference) + ')')  # 回退历史记录
            # if node["parameters"]["historyWait"] > 2:  # 回退后要等待的时间
            time.sleep(node["parameters"]["historyWait"])
            # else:
            # time.sleep(2)
            try:
                self.browser.execute_script('window.stop()')
            except:
                pass
        ti = 0
        # print("CURRENT URL:", self.browser.current_url)
        # time.sleep(2)
        # if self.browser.current_url.startswith("data:") or self.browser.current_url.startswith("chrome:"):
        if self.browser.current_url != thisHistoryURL and self.history["index"] != thisHistoryLength and self.history["handle"] == self.browser.current_window_handle:
            while self.browser.current_url != thisHistoryURL:  # 如果执行完一次循环之后网址发生了变化
                try:
                    self.browser.execute_script("history.go(1)")  # 如果是data:开头的网址，就前进一步
                except:  # 超时的情况下
                    pass
                ti += 1
                if self.browser.current_url == thisHistoryURL or ti > thisHistoryLength:  # 如果执行完一次循环之后网址发生了变化
                    break
            time.sleep(2)
            if xpath != "":
                if element == None: # 不固定元素列表
                    element = self.browser.find_elements(By.XPATH, xpath, iframe=node["parameters"]["iframe"])
                else: # 固定元素列表
                    element = self.browser.find_element(By.XPATH, xpath, iframe=node["parameters"]["iframe"])
                # if index > 0:
                    # index -= 1  # 如果是data:开头的网址，就要重试一次
        else:
            if element == None:
                element = elements
        return index, element

    # 对循环的处理
    def loopExecute(self, node, loopValue, loopPath="", index=0):
        if self.shutdown_event.is_set():
            return
        self.event.wait()
        time.sleep(0.1)  # 第一次执行循环的时候强制等待1秒
        thisHandle = self.browser.current_window_handle  # 记录本次循环内的标签页的ID
        try:
            thisHistoryLength = self.browser.execute_script(
                'return history.length')  # 记录本次循环内的history的length
        except:
            thisHistoryLength = 0
        self.history["index"] = thisHistoryLength
        self.history["handle"] = thisHandle
        thisHistoryURL = self.browser.current_url
        # 快速提取处理
        # start = time.time()
        try:
            tree = html.fromstring(self.browser.page_source)
        except Exception as e:
            self.print_and_log("解析页面时出错，将切换普通提取模式|Error parsing page, will switch to normal extraction mode")
            node["parameters"]["quickExtractable"] = False
        # end = time.time()
        # print("解析页面秒数：", end - start)
        if node["parameters"]["quickExtractable"]:
            self.browser.switch_to.default_content() # 切换到主页面
            tree = html.fromstring(self.browser.page_source)
            if int(node["parameters"]["loopType"]) == 1: # 不固定元素列表
                baseXPath = replace_field_values(node["parameters"]["baseXPath"], self.outputParameters, self)
                rows = tree.xpath(baseXPath)
            elif int(node["parameters"]["loopType"]) == 2: # 固定元素列表
                rows = []
                for path in node["parameters"]["baseXPath"].split("\n"):
                    baseXPath = replace_field_values(path, self.outputParameters, self)
                    rows.extend(tree.xpath(baseXPath))
                
            for row in rows:
                if node["parameters"]["clear"] == 1:
                    self.clearOutputParameters()
                for param in node["parameters"]["quickParams"]:
                    xpath = replace_field_values(param["xpath"], self.outputParameters, self)
                    content = row.xpath(xpath)
                    try:
                        content = ' '.join(result.strip()
                                        for result in content if result.strip())
                        # 链接或者图片的情况下，合并链接相对路径为绝对路径
                        if param["nodeType"] == 2 or param["nodeType"] == 4:
                            base_url = self.browser.current_url
                            # 合并链接相对路径为绝对路径
                            content = urljoin(base_url, content)
                        if len(content) == 0:
                            content = param["default"]
                    except:
                        content = param["default"]
                    self.outputParameters[param["name"]] = content
                if node["parameters"]["newLine"]:
                    line = new_line(self.outputParameters,
                            self.maxViewLength, self.outputParametersRecord)
                    self.OUTPUT.append(line)
            self.saveData()
        elif int(node["parameters"]["loopType"]) == 0:  # 单个元素循环
            # 无跳转标签页操作
            count = 0  # 执行次数
            bodyText = "-"

            while True:  # do while循环
                if self.shutdown_event.is_set():
                    break
                try:
                    finished = False
                    if node["parameters"]["exitCount"] == 0:
                        # newBodyText = self.browser.find_element(By.XPATH, node["parameters"]["exitElement"], iframe=node["parameters"]["iframe"]).text
                        # 用find_elements获取所有匹配到的文本
                        try:
                            exitElements = self.browser.find_elements(By.XPATH, node["parameters"]["exitElement"], iframe=node["parameters"]["iframe"])
                            newBodyText = ""
                            for exitElement in exitElements:
                                newBodyText += exitElement.text
                        except Exception as e:
                            self.print_and_log(f"设定的退出循环元素：{node['parameters']['exitElement']}的文本无法获取，本次循环将不再检测元素文本是否变化，将会继续执行，为解决此问题，您可以修改检测元素文本不变的元素为其他元素，或者将循环次数设定为固定次数大于0的值。")
                            self.print_and_log(f"The text of the exit loop element set: {node['parameters']['exitElement']} cannot be obtained, this loop will no longer check whether the text of the element has changed, and will continue to execute. To solve this problem, you can modify the element whose text does not change to other elements, or set the number of loops to a fixed number greater than 0.")
                            self.print_and_log(e)
                            exitElements = []
                            # newBodyText为随机文本，保证一直执行
                            newBodyText = str(random.random())
                        if node["parameters"]["iframe"]:  # 如果标记了iframe
                            iframes = self.browser.find_elements(
                                By.CSS_SELECTOR, "iframe", iframe=False)
                            for iframe in iframes:
                                self.browser.switch_to.default_content()
                                self.browser.switch_to.frame(iframe)
                                iframe_text = super(self.browser.__class__, self.browser).find_element(
                                    By.CSS_SELECTOR, "body").text  # 用super调用父类的方法
                                newBodyText += iframe_text
                                self.browser.switch_to.default_content()
                        if newBodyText == bodyText:  # 如果页面内容无变化
                            self.print_and_log("页面已检测不到新内容，停止循环。")
                            self.print_and_log(
                                "No new content detected on the page, stop loop.")
                            finished = True
                            break
                        else:
                            self.print_and_log("检测到页面变化，继续循环。")
                            self.print_and_log(
                                "Page changed detected, continue loop.")
                            bodyText = newBodyText
                    xpath = replace_field_values(
                        node["parameters"]["xpath"], self.outputParameters, self)
                    # self.recordLog("循环元素|Loop element:", xpath)
                    element = self.browser.find_element(
                        By.XPATH, xpath, iframe=node["parameters"]["iframe"])
                    for i in node["sequence"]:  # 挨个执行操作
                        self.executeNode(
                            i, element, xpath, 0)
                        if self.BREAK or self.CONTINUE:  # 如果有break操作，下面的操作不执行
                            self.CONTINUE = False
                            break
                    if self.BREAK:  # 如果有break操作，退出循环
                        self.BREAK = False
                        finished = True
                        break
                    finished = True
                except NoSuchElementException:
                    # except:
                    self.print_and_log("Single loop element not found: ",
                                       xpath)
                    self.print_and_log("找不到要循环的单个元素: ", xpath)
                    for i in node["sequence"]:  # 不带点击元素的把剩余的如提取数据的操作执行一遍
                        if node["option"] != 2:
                            self.executeNode(
                                i, None, xpath, 0)
                    finished = True
                    break  # 如果找不到元素，退出循环
                finally:
                    if not finished:
                        self.print_and_log("\n\n-------Retrying-------\n\n")
                        self.print_and_log("-------Retrying-------: ",
                                           node["parameters"]["xpath"])
                        for i in node["sequence"]:  # 不带点击元素的把剩余的如提取数据的操作执行一遍
                            if node["option"] != 2:
                                self.executeNode(
                                    i, None, xpath, 0)
                        break  # 如果找不到元素，退出循环
                count = count + 1
                self.print_and_log("Page: ", count)
                # self.print_and_log(node["parameters"]["exitCount"], "-------")
                if node["parameters"]["exitCount"] == count:  # 如果达到设置的退出循环条件的话
                    break
                if int(node["parameters"]["breakMode"]) > 0:  # 如果设置了退出循环的脚本条件
                    output = self.execute_code(int(
                        node["parameters"]["breakMode"]) - 1, node["parameters"]["breakCode"],
                                               node["parameters"]["breakCodeWaitTime"],
                                               iframe=node["parameters"]["iframe"])
                    code = get_output_code(output)
                    if code <= 0:
                        break
        elif int(node["parameters"]["loopType"]) == 1:  # 不固定元素列表
            try:
                xpath = replace_field_values(
                    node["parameters"]["xpath"], self.outputParameters, self)
                elements = self.browser.find_elements(By.XPATH,
                                                      xpath, iframe=node["parameters"]["iframe"])
                # self.recordLog("循环元素|Loop element:", xpath)
                if len(elements) == 0:
                    self.print_and_log("Loop element not found: ",
                                       xpath)
                    self.print_and_log("找不到循环元素：", xpath)
                index = 0
                skipCount = node["parameters"]["skipCount"]
                while index < len(elements):
                    if self.shutdown_event.is_set():
                        break
                    if index < skipCount:
                        index += 1
                        self.print_and_log("跳过第" + str(index) + "个元素")
                        self.print_and_log("Skip the " + str(index) + "th element")
                        continue
                    try:
                        element = elements[index]
                        element_text = element.text
                    except StaleElementReferenceException: # 如果元素已经失效，重试
                        self.print_and_log("元素已失效，重新获取元素|Element has expired, reacquiring element")
                        elements = self.browser.find_elements(By.XPATH,
                                                              xpath, iframe=node["parameters"]["iframe"])
                        element = elements[index]
                    for i in node["sequence"]:  # 挨个顺序执行循环里所有的操作
                        self.executeNode(i, element,
                                         xpath, index)
                        if self.BREAK or self.CONTINUE:  # 如果有break操作，下面的操作不执行
                            self.CONTINUE = False
                            break
                    if self.BREAK:
                        self.BREAK = False
                        break
                    index, elements = self.handleHistory(node, xpath, thisHandle, thisHistoryURL, thisHistoryLength, index, elements=elements)
                    if int(node["parameters"]["breakMode"]) > 0:  # 如果设置了退出循环的脚本条件
                        output = self.execute_code(int(
                            node["parameters"]["breakMode"]) - 1, node["parameters"]["breakCode"],
                                                   node["parameters"]["breakCodeWaitTime"],
                                                   iframe=node["parameters"]["iframe"])
                        code = get_output_code(output)
                        if code <= 0:
                            break
                    index = index + 1
            except NoSuchElementException:
                self.print_and_log("Loop element not found: ", xpath)
                self.print_and_log("找不到循环元素：", xpath)
            except Exception as e:
                raise
        elif int(node["parameters"]["loopType"]) == 2:  # 固定元素列表
            # 千万不要忘了分割！！
            paths = node["parameters"]["pathList"].split("\n")
            # for path in node["parameters"]["pathList"].split("\n"):
            index = 0
            skipCount = node["parameters"]["skipCount"]
            while index < len(paths):
                if self.shutdown_event.is_set():
                    break
                if index < skipCount:
                    index += 1
                    self.print_and_log("跳过第" + str(index) + "个元素")
                    self.print_and_log("Skip the " + str(index) + "th element")
                    continue
                path = paths[index]
                try:
                    path = replace_field_values(
                        path, self.outputParameters, self)
                    element = self.browser.find_element(
                        By.XPATH, path, iframe=node["parameters"]["iframe"])
                    # self.recordLog("循环元素|Loop element:", path)
                    for i in node["sequence"]:  # 挨个执行操作
                        self.executeNode(i, element, path, 0)
                        if self.BREAK or self.CONTINUE:  # 如果有break操作，下面的操作不执行
                            self.CONTINUE = False
                            break
                    if self.BREAK:
                        self.BREAK = False
                        break
                    index, element = self.handleHistory(node, path, thisHandle, thisHistoryURL, thisHistoryLength, index, element=element)
                except NoSuchElementException:
                    self.print_and_log("Loop element not found: ", path)
                    self.print_and_log("找不到循环元素：", path)
                    index += 1
                    continue  # 循环中找不到元素就略过操作
                except Exception as e:
                    raise
                if int(node["parameters"]["breakMode"]) > 0:  # 如果设置了退出循环的脚本条件
                    output = self.execute_code(int(
                        node["parameters"]["breakMode"]) - 1, node["parameters"]["breakCode"],
                                               node["parameters"]["breakCodeWaitTime"],
                                               iframe=node["parameters"]["iframe"])
                    code = get_output_code(output)
                    if code <= 0:
                        break
                index = index + 1
        elif int(node["parameters"]["loopType"]) == 3:  # 固定文本列表
            textList = node["parameters"]["textList"].split("\n")
            if len(textList) == 1:  # 如果固定文本列表只有一行，现在就可以替换变量
                textList = replace_field_values(
                    node["parameters"]["textList"], self.outputParameters, self).split("\n")
            skipCount = node["parameters"]["skipCount"]
            index = 0
            for text in textList:
                if self.shutdown_event.is_set():
                    break
                if index < skipCount:
                    index += 1
                    self.print_and_log("跳过第" + str(index) + "个文本")
                    self.print_and_log("Skip the " + str(index) + "th text")
                    continue
                text = replace_field_values(text, self.outputParameters, self)
                # self.recordLog("当前循环文本|Current loop text:", text)
                for i in node["sequence"]:  # 挨个执行操作
                    self.executeNode(i, text, "", 0)
                    if self.BREAK or self.CONTINUE:  # 如果有break操作，下面的操作不执行
                        self.CONTINUE = False
                        break
                if self.BREAK:
                    self.BREAK = False
                    break
                if int(node["parameters"]["breakMode"]) > 0:  # 如果设置了退出循环的脚本条件
                    output = self.execute_code(int(
                        node["parameters"]["breakMode"]) - 1, node["parameters"]["breakCode"],
                                               node["parameters"]["breakCodeWaitTime"],
                                               iframe=node["parameters"]["iframe"])
                    code = get_output_code(output)
                    if code <= 0:
                        break
                index, _ = self.handleHistory(node, "", thisHandle, thisHistoryURL, thisHistoryLength, index)
        elif int(node["parameters"]["loopType"]) == 4:  # 固定网址列表
            # tempList = node["parameters"]["textList"].split("\r\n")
            urlList = list(
                filter(isnotnull, node["parameters"]["textList"].split("\n")))  # 去空行
            if len(urlList) == 1:  # 如果固定网址列表只有一行，现在就可以替换变量
                urlList = replace_field_values(
                    node["parameters"]["textList"], self.outputParameters, self).split("\n")
            skipCount = node["parameters"]["skipCount"]
            index = 0
            for url in urlList:
                if self.shutdown_event.is_set():
                    break
                if index < skipCount:
                    index += 1
                    self.print_and_log("跳过第" + str(index) + "个网址")
                    self.print_and_log("Skip the " + str(index) + "th url")
                    continue
                url = replace_field_values(url, self.outputParameters, self)
                # self.recordLog("当前循环网址|Current loop url:", url)
                for i in node["sequence"]:
                    self.executeNode(i, url, "", 0)
                    if self.BREAK or self.CONTINUE:  # 如果有break操作，下面的操作不执行
                        self.CONTINUE = False
                        break
                if self.BREAK:
                    self.BREAK = False
                    break
                if int(node["parameters"]["breakMode"]) > 0:  # 如果设置了退出循环的脚本条件
                    output = self.execute_code(int(
                        node["parameters"]["breakMode"]) - 1, node["parameters"]["breakCode"],
                                               node["parameters"]["breakCodeWaitTime"],
                                               iframe=node["parameters"]["iframe"])
                    code = get_output_code(output)
                    if code <= 0:
                        break
        elif int(node["parameters"]["loopType"]) <= 7:  # 命令返回值
            while True:  # do while循环
                if self.shutdown_event.is_set():
                    break
                if int(node["parameters"]["loopType"]) == 5:  # JS
                    output = self.execute_code(
                        0, node["parameters"]["code"], node["parameters"]["waitTime"],
                        iframe=node["parameters"]["iframe"])
                elif int(node["parameters"]["loopType"]) == 6:  # System
                    output = self.execute_code(
                        1, node["parameters"]["code"], node["parameters"]["waitTime"],
                        iframe=node["parameters"]["iframe"])
                elif int(node["parameters"]["loopType"]) == 7:  # Python
                    output = self.execute_code(
                        6, node["parameters"]["code"], node["parameters"]["waitTime"],
                        iframe=node["parameters"]["iframe"])
                code = get_output_code(output)
                if code <= 0:
                    break
                for i in node["sequence"]:  # 挨个执行操作
                    self.executeNode(i, code, node["parameters"]["xpath"], 0)
                    if self.BREAK or self.CONTINUE:  # 如果有break操作，下面的操作不执行
                        self.CONTINUE = False
                        break
                if self.BREAK:
                    self.BREAK = False
                    break
        self.history["index"] = thisHistoryLength
        self.history["handle"] = self.browser.current_window_handle
        self.scrollDown(node["parameters"])

    # 打开网页操作
    def openPage(self, param, loopValue):
        if self.shutdown_event.is_set():
            return
        self.event.wait()
        time.sleep(1)  # 打开网页后强行等待至少1秒
        if len(self.browser.window_handles) > 1:
            self.browser.switch_to.window(
                self.browser.window_handles[-1])  # 打开网页操作从第1个页面开始
            try:
                self.browser.close()
            except:
                pass
        self.browser.switch_to.window(
            self.browser.window_handles[0])  # 打开网页操作从第1个页面开始
        self.history["handle"] = self.browser.current_window_handle
        if param["useLoop"]:
            url = loopValue
        elif param["url"] != "about:blank":
            url = self.links[self.urlId]
            # clear output parameters
            for key in self.outputParameters:
                self.outputParameters[key] = ""
        else:  # 在流程图其他位置设置了打开网页的操作，读取的应该是第一个网址，如打开网页后登录，再打开第二个网页
            url = list(filter(isnotnull, param["links"].split("\n")))[0]
        # 将value中的Field[""]替换为outputParameters中的键值
        url = replace_field_values(url, self.outputParameters, self)
        try:
            maxWaitTime = int(param["maxWaitTime"])
        except:
            maxWaitTime = 10  # 默认最大等待时间为10秒
        try:
            self.browser.set_page_load_timeout(maxWaitTime)  # 加载页面最大超时时间
            self.browser.set_script_timeout(maxWaitTime)
            self.browser.get(url)
            if param["cookies"] != "":
                self.browser.delete_all_cookies()  # 清除所有已有cookie
                cookies = param["cookies"].split('\n')
                for cookie in cookies:
                    name, value = cookie.split('=', 1)
                    cookie_dict = {'name': name, 'value': value}
                    # 加载 cookie
                    self.browser.add_cookie(cookie_dict)
            self.print_and_log('加载页面|Loading page: ' + url)
        except TimeoutException:
            self.print_and_log(
                'Time out after set seconds when loading page: ' + url)
            try:
                self.browser.execute_script('window.stop()')
            except:
                pass
        except Exception as e:
            self.print_and_log("Failed to load page: " + url)
        try:
            self.history["index"] = self.browser.execute_script(
                "return history.length")
        except TimeoutException:
            try:
                self.browser.execute_script('window.stop()')
                self.history["index"] = self.browser.execute_script(
                    "return history.length")
            except:
                self.history["index"] = 0
        except Exception as e:
            self.print_and_log("History Length Error")
            self.history["index"] = 0
        self.scrollDown(param)  # 控制屏幕向下滚动

    # 键盘输入操作
    def inputInfo(self, param, loopValue):
        if self.shutdown_event.is_set():
            return
        self.event.wait()
        time.sleep(0.1)  # 输入之前等待0.1秒
        try:
            xpath = replace_field_values(
                param["xpath"], self.outputParameters, self)
            textbox = self.browser.find_element(
                By.XPATH, xpath, iframe=param["iframe"])
            #     textbox.send_keys(Keys.CONTROL, 'a')
            #     textbox.send_keys(Keys.BACKSPACE)
            self.execute_code(
                2, param["beforeJS"], param["beforeJSWaitTime"], textbox, iframe=param["iframe"])  # 执行前置JS
            # Send the HOME key
            textbox.send_keys(Keys.HOME)
            # Send the SHIFT + END key combination
            textbox.send_keys(Keys.SHIFT, Keys.END)
            # Send the DELETE key
            textbox.send_keys(Keys.DELETE)
            value = ""
            if param["useLoop"]:
                value = loopValue
            else:
                value = param["value"]
            # 将value中的Field[""]替换为outputParameters中的键值
            # pattern = r'Field\["([^"]+)"\]'
            try:
                # replaced_text = re.sub(
                    # pattern, lambda match: self.outputParameters.get(match.group(1), ''), value)
                replaced_text = replace_field_values(value, self.outputParameters, self)
                replaced_text = re.sub(
                    '<enter>', '', replaced_text, flags=re.IGNORECASE)
            except:
                replaced_text = value
            index = param["index"]
            if index != 0:
                try:
                    replaced_text = replaced_text.split("~")[index - 1]
                except:
                    self.print_and_log("取值失败，可能是因为取值索引超出范围，将使用整个文本值")
                    self.print_and_log(
                        "Failed to get value, maybe because the index is out of range, will use the entire text value")
            textbox.send_keys(replaced_text)
            if value.lower().find("<enter>") >= 0:
                textbox.send_keys(Keys.ENTER)
            self.recordLog("输入文字|Input text: " +
                           replaced_text + " to " + xpath)
            self.execute_code(
                2, param["afterJS"], param["afterJSWaitTime"], textbox, iframe=param["iframe"])  # 执行后置js
        except:
            self.print_and_log("Cannot find input box element:" +
                               xpath + ", please try to set the wait time before executing this operation")
            self.print_and_log("找不到输入框元素:" + xpath + "，请尝试在执行此操作前设置等待时间")

    # 点击元素操作
    def clickElement(self, param, loopElement=None, clickPath="", index=0):
        if self.shutdown_event.is_set():
            return
        self.event.wait()
        try:
            maxWaitTime = int(param["maxWaitTime"])
        except:
            maxWaitTime = 10
        self.browser.set_page_load_timeout(maxWaitTime)  # 加载页面最大超时时间
        self.browser.set_script_timeout(maxWaitTime)
        # 点击前对该元素执行一段JavaScript代码
        try:
            # element = self.browser.find_element(
            #     By.XPATH, path, iframe=param["iframe"])
            clickPath = replace_field_values(
                clickPath, self.outputParameters, self)
            xpath = replace_field_values(
                param["xpath"], self.outputParameters, self)
            if xpath.find("point(") >= 0:  # 如果xpath中包含point()，说明是相对坐标的点击
                index = 0
                path = "//body"
            elif param["useLoop"]:  # 使用循环的情况下，传入的clickPath就是实际的xpath
                if xpath == "":
                    path = clickPath
                else:
                    path = "(" + clickPath + ")" + \
                           "[" + str(index + 1) + "]" + \
                           xpath
                    index = 0  # 如果是相对循环内元素的点击，在定位到元素后，index应该重置为0
                # element = loopElement
            else:
                index = 0
                path = xpath  # 不然使用元素定义的xpath
                # element = self.browser.find_element(
                #     By.XPATH, path, iframe=param["iframe"])
            elements = self.browser.find_elements(
                By.XPATH, path, iframe=param["iframe"])
            element = elements[index]
            if param["beforeJS"] != "":
                self.execute_code(2, param["beforeJS"],
                                  param["beforeJSWaitTime"], element, iframe=param["iframe"])
        except:
            self.print_and_log("Cannot find element:" +
                               path + ", please try to set the wait time before executing this operation")
            self.print_and_log("找不到要点击的元素:" + path + "，请尝试在执行此操作前设置等待时间")
        tempHandleNum = len(self.browser.window_handles)  # 记录之前的窗口位置
        try:
            click_way = int(param["clickWay"])
        except:
            click_way = 0
        try:
            newTab = int(param["newTab"])
        except:
            newTab = 0
        try:
            if xpath.find("point(") >= 0:  # 如果xpath中包含point()，说明是相对坐标的点击
                point = xpath.split("point(")[1].split(")")[0].split(",")
                x = int(point[0])
                y = int(point[1])
                # try:
                #     actions = ActionChains(self.browser)  # 实例化一个action对象
                #     actions.move_to_element(element).perform()
                #     actions.move_by_offset(x, y).perform()
                #     actions.click().perform()
                # except Exception as e:
                script = "document.elementFromPoint(" + str(x) + "," + str(y) + ").click();"
                self.browser.execute_script(script)
            elif click_way == 0:  # 用selenium的点击方法
                try:
                    actions = ActionChains(self.browser)  # 实例化一个action对象
                    if newTab == 1:  # 在新标签页打开
                        if sys.platform == "darwin":  # Mac
                            actions.key_down(Keys.COMMAND).click(element).key_up(Keys.COMMAND).perform()
                        else:
                            # Ctrl + Click
                            actions.key_down(Keys.CONTROL).click(element).key_up(Keys.CONTROL).perform()
                    else:
                        actions.click(element).perform()
                except Exception as e:
                    self.browser.execute_script("arguments[0].scrollIntoView();", element)
                    try:
                        actions = ActionChains(self.browser)  # 实例化一个action对象
                        actions.click(element).perform()
                    except Exception as e:
                        self.print_and_log(f"Selenium点击元素{path}失败，将尝试使用JavaScript点击")
                        self.print_and_log(f"Failed to click element {path} with Selenium, will try to click with JavaScript")
                        script = 'var result = document.evaluate(`' + path + \
                            '`, document, null, XPathResult.ANY_TYPE, null);for(let i=0;i<arguments[0];i++){result.iterateNext();} result.iterateNext().click();'
                        self.browser.execute_script(script, str(index))  # 用js的点击方法
            elif click_way == 1:  # 用js的点击方法
                script = 'var result = document.evaluate(`' + path + \
                         '`, document, null, XPathResult.ANY_TYPE, null);for(let i=0;i<arguments[0];i++){result.iterateNext();} result.iterateNext().click();'
                self.browser.execute_script(script, str(index))  # 用js的点击方法
            elif click_way == 2: # 双击
                try:
                    actions = ActionChains(self.browser)  # 实例化一个action对象
                    actions.double_click(element).perform()
                except Exception as e:
                    self.browser.execute_script("arguments[0].scrollIntoView();", element)
                    try:
                        actions = ActionChains(self.browser)  # 实例化一个action对象
                        actions.double_click(element).perform()
                    except Exception as e:
                        self.print_and_log(f"Selenium双击元素{path}失败，将尝试使用JavaScript双击")
                        self.print_and_log(f"Failed to double click element {path} with Selenium, will try to double click with JavaScript")
                        script = 'var result = document.evaluate(`' + path + \
                            '`, document, null, XPathResult.ANY_TYPE, null);for(let i=0;i<arguments[0];i++){result.iterateNext();} result.iterateNext().click();'
                        self.browser.execute_script(script, str(index))  # 用js的点击方法
            self.recordLog("点击元素|Click element: " + path)
        except TimeoutException:
            self.print_and_log(
                'Time out after set seconds when loading clicked page')
            try:
                self.browser.execute_script('window.stop()')
            except:
                pass
        except Exception as e:
            self.print_and_log(
                "点击元素失败:" + path, "，请尝试将点击类型改为JavaScript点击后重试。")
            self.print_and_log("Failed to click element:" + path,
                               ", please try to change the click type to JavaScript Click.")
            self.print_and_log(e)

        # 弹窗处理
        if param["alertHandleType"] > 0:
            try:
                time.sleep(1.5)
                alert = self.browser.switch_to.alert
                alertHandleType = int(param["alertHandleType"])
                if alertHandleType == 1:
                    alert.accept()
                    self.print_and_log("已点击确认|Clicked OK")
                elif alertHandleType == 2:
                    alert.dismiss()
                    self.print_and_log("已点击取消|Clicked Cancel")
            except Exception as e:
                self.print_and_log("找不到弹窗|Cannot find alert")

        # 点击后对该元素执行一段JavaScript代码
        try:
            if param["afterJS"] != "":
                element = self.browser.find_element(
                    By.XPATH, path, iframe=param["iframe"])
                self.execute_code(2, param["afterJS"],
                                  param["afterJSWaitTime"], element, iframe=param["iframe"])
        except:
            self.print_and_log("Cannot find element:" + path)
            self.print_and_log("找不到要点击的元素:" + path + "，请尝试在执行此操作前设置等待时间")
        waitTime = float(param["wait"]) + 0.01  # 点击之后等待
        try:
            waitType = int(param["waitType"])
        except:
            waitType = 0
        if waitType == 0:  # 固定等待时间
            time.sleep(waitTime)
        elif waitType == 1:  # 随机等待时间
            time.sleep(random.uniform(waitTime * 0.5, waitTime * 1.5))
        if tempHandleNum != len(self.browser.window_handles):  # 如果有新标签页的行为发生
            self.browser.switch_to.window(
                self.browser.window_handles[-1])  # 跳转到新的标签页
            self.history["handle"] = self.browser.current_window_handle
            try:
                self.history["index"] = self.browser.execute_script(
                    "return history.length")
            except TimeoutException:
                try:
                    self.browser.execute_script('window.stop()')
                except:
                    pass
                self.history["index"] = self.browser.execute_script(
                    "return history.length")
            except Exception as e:
                self.print_and_log("History Length Error")
                self.history["index"] = 0
        else:
            try:
                self.history["index"] = self.browser.execute_script(
                    "return history.length")
            except TimeoutException:
                try:
                    self.browser.execute_script('window.stop()')
                except:
                    pass
                self.history["index"] = self.browser.execute_script(
                    "return history.length")
                # 如果打开了新窗口，切换到新窗口
            except Exception as e:
                self.print_and_log("History Length Error")
                self.history["index"] = 0
        self.scrollDown(param)  # 根据参数配置向下滚动

    def get_content(self, p, element):
        if self.shutdown_event.is_set():
            return ""
        self.event.wait()
        # self.print_and_log(p)
        content = ""
        if p["contentType"] == 0:
            # 先处理特殊节点类型
            if p["nodeType"] == 2:
                if element.get_attribute("href") != None:
                    content = element.get_attribute("href")
                else:
                    content = ""
            elif p["nodeType"] == 3:
                if element.get_attribute("value") != None:
                    content = element.get_attribute("value")
                else:
                    content = ""
            elif p["nodeType"] == 4:  # 图片
                if element.get_attribute("src") != None:
                    content = element.get_attribute("src")
                else:
                    content = ""
                try:
                    downloadPic = p["downloadPic"]
                except:
                    downloadPic = 0
                if downloadPic == 1:
                    download_image(self, content, "Data/Task_" +
                                   str(self.id) + "/" + self.saveName + "/images", element)
            else:  # 普通节点
                if p["splitLine"] == 1:
                    text = extract_text_from_html(element.get_attribute('outerHTML'))
                    content = split_text_by_lines(text)
                else:
                    content = element.text
        elif p["contentType"] == 1:  # 只采集当期元素下的文本，不包括子元素
            if p["nodeType"] == 2:
                if element.get_attribute("href") != None:
                    content = element.get_attribute("href")
                else:
                    content = ""
            elif p["nodeType"] == 3:
                if element.get_attribute("value") != None:
                    content = element.get_attribute("value")
                else:
                    content = ""
            elif p["nodeType"] == 4:  # 图片
                if element.get_attribute("src") != None:
                    content = element.get_attribute("src")
                else:
                    content = ""
                try:
                    downloadPic = p["downloadPic"]
                except:
                    downloadPic = 0
                if downloadPic == 1:
                    download_image(self, content, "Data/Task_" +
                                   str(self.id) + "/" + self.saveName + "/images", element)
            else:
                command = 'var arr = [];\
                var content = arguments[0];\
                for(var i = 0, len = content.childNodes.length; i < len; i++) {\
                    if(content.childNodes[i].nodeType === 3){  \
                        arr.push(content.childNodes[i].nodeValue);\
                    }\
                }\
                var str = arr.join(" "); \
                return str;'
                content = self.browser.execute_script(command, element).replace(
                    "\n", "").replace("\\s+", " ")
        elif p["contentType"] == 2:
            content = element.get_attribute('innerHTML')
        elif p["contentType"] == 3:
            content = element.get_attribute('outerHTML')
        elif p["contentType"] == 4:
            # 获取元素的背景图片地址
            bg_url = element.value_of_css_property('background-image')
            # 清除背景图片地址中的多余字符
            bg_url = bg_url.replace('url("', '').replace('")', '')
            content = bg_url
        elif p["contentType"] == 5:
            content = self.browser.current_url
        elif p["contentType"] == 6:
            content = self.browser.title
        elif p["contentType"] == 7:
            # 获取整个网页的高度和宽度
            size = self.browser.get_window_size()
            width = size["width"]
            height = size["height"]
            # 调整浏览器窗口的大小
            if self.commandline_config["headless"] == 1: # 无头模式下，截取整个网页的高度
                page_width = self.browser.execute_script(
                    "return document.body.scrollWidth")
                page_height = self.browser.execute_script(
                    "return document.body.scrollHeight")
                self.browser.set_window_size(page_width, page_height)
                time.sleep(1)
            else:
                self.browser.set_window_size(width, height)
            element.screenshot("Data/Task_" + str(self.id) + "/" + self.saveName +
                               "/screenshots/" + str(time.time()) + ".png")
            # 截图完成后，将浏览器的窗口大小设置为原来的大小
            self.browser.set_window_size(width, height)
        elif p["contentType"] == 8:
            try:
                size = self.browser.get_window_size()
                width = size["width"]
                height = size["height"]
                screenshot = element.screenshot_as_png
                screenshot_stream = io.BytesIO(screenshot)
                # 使用Pillow库打开截图，并转换为灰度图像
                image = Image.open(screenshot_stream).convert('L')
                temp_name = "OCR_" + str(time.time()) + ".png"
                location = "Data/Task_" + \
                           str(self.id) + "/" + self.saveName + "/" + temp_name
                image.save(location)
                ocr = DdddOcr(show_ad=False)
                with open(location, 'rb') as f:
                    image_bytes = f.read()
                content = ocr.classification(image_bytes)
                os.remove(location)
                self.browser.set_window_size(width, height)
                # 使用Tesseract OCR引擎识别图像中的文本
                # content = pytesseract.image_to_string(image,  lang='chi_sim+eng')
            except Exception as e:
                # try:
                #     self.print_and_log(e)
                #     self.print_and_log("识别中文失败，尝试只识别英文")
                #     self.print_and_log("Failed to recognize Chinese, try to recognize English only")
                #     screenshot = element.screenshot_as_png
                #     screenshot_stream = io.BytesIO(screenshot)
                #     # 使用Pillow库打开截图，并转换为灰度图像
                #     image = Image.open(screenshot_stream).convert('L')
                #     # 使用Tesseract OCR引擎识别图像中的文本
                #     # content = pytesseract.image_to_string(image,  lang='eng')
                # except Exception as e:
                content = "OCR Error"
                self.print_and_log(e)
                # if sys.platform == "win32":
                #     self.print_and_log("要使用OCR识别功能，你需要安装Tesseract-OCR并将其添加到环境变量PATH中（添加后需重启EasySpider）：https://blog.csdn.net/u010454030/article/details/80515501")
                #     self.print_and_log("\nhttps://www.bilibili.com/video/BV1GP411y7u4/")
                # elif sys.platform == "darwin":
                #     self.print_and_log(
                #         "注意以上错误，要使用OCR识别功能，你需要安装Tesseract-OCR并将其添加到环境变量PATH中（添加后需重启EasySpider）：https://zhuanlan.zhihu.com/p/146044810")
                # elif sys.platform == "linux":
                #     self.print_and_log(
                #         "注意以上错误，要使用OCR识别功能，你需要安装Tesseract-OCR并将其添加到环境变量PATH中（添加后需重启EasySpider）：https://zhuanlan.zhihu.com/p/420259031")
                # else:
                #     self.print_and_log("注意以上错误，要使用OCR识别功能，你需要安装Tesseract-OCR并将其添加到环境变量PATH中（添加后需重启EasySpider）：https://blog.csdn.net/u010454030/article/details/80515501")
                #     self.print_and_log("\nhttps://www.bilibili.com/video/BV1GP411y7u4/")
                # self.print_and_log("To use OCR, You need to install Tesseract-OCR and add it to the environment variable PATH (need to restart EasySpider after you put in PATH): https://tesseract-ocr.github.io/tessdoc/Installation.html")
        elif p["contentType"] == 9:
            content = self.execute_code(
                2, p["JS"], p["JSWaitTime"], element, iframe=p["iframe"])
        elif p["contentType"] == 12:  # 系统命令返回值
            content = self.execute_code(1, p["JS"], p["JSWaitTime"])
        elif p["contentType"] == 13:  # eval返回值
            content = self.execute_code(6, p["JS"], p["JSWaitTime"])
        elif p["contentType"] == 10:  # 下拉框选中的值
            try:
                select_element = Select(element)
                content = select_element.first_selected_option.get_attribute(
                    "value")
            except:
                content = ""
        elif p["contentType"] == 11:  # 下拉框选中的文本
            try:
                select_element = Select(element)
                content = select_element.first_selected_option.text
            except:
                content = ""
        elif p["contentType"] == 14:  # 元素属性值
            attribute_name = p["JS"]
            try:
                content = element.get_attribute(attribute_name)
            except:
                content = ""
        elif p["contentType"] == 15:  # 常量值
            content = p["JS"]
        if content == None:
            content = ""
        return content

    def clearOutputParameters(self):
        if self.shutdown_event.is_set():
            return
        self.event.wait()

        for key in self.outputParameters:
            self.outputParameters[key] = ""
        self.recordLog("清空输出参数|Clear output parameters")

    # 提取数据操作
    def getData(self, param, loopElement, isInLoop=True, parentPath="", index=0):
        if self.shutdown_event.is_set():
            return
        self.event.wait()
        parentPath = replace_field_values(
            parentPath, self.outputParameters, self)
        if param["clear"] == 1:
            self.clearOutputParameters()
        try:
            pageHTML = etree.HTML(self.browser.page_source)
        except:
            pageHTML = etree.HTML("")
        if loopElement != "":  # 只在数据在循环中提取时才需要获取循环元素
            try:
                loopElementOuterHTML = loopElement.get_attribute('outerHTML')
            except:
                try:  # 循环点击每个链接如果没有新标签页打开，loopElement会丢失，此时需要重新获取
                    elements = self.browser.find_elements(
                        By.XPATH, parentPath, iframe=param["params"][0]["iframe"])
                    loopElement = elements[index]
                    loopElementOuterHTML = loopElement.get_attribute(
                        'outerHTML')
                except:
                    loopElementOuterHTML = ""
        else:
            loopElementOuterHTML = ""
        loopElementHTML = etree.HTML(loopElementOuterHTML)
        for p in param["params"]:
            if p["optimizable"]:
                try:
                    relativeXPath = replace_field_values(
                        p["relativeXPath"], self.outputParameters, self)
                    # 只有当前环境不变变化才可以快速提取数据
                    if self.browser.iframe_env != p["iframe"]:
                    # if p["iframe"] or self.browser.iframe_env != p["iframe"]: # 如果是iframe，则不能快速提取数据，主要是各个上下文的iframe切换，但一般不会有人这么做
                        p["optimizable"] = False
                        continue
                    # relativeXPath = relativeXPath.lower()
                    # relativeXPath = lowercase_tags_in_xpath(relativeXPath)
                    # 已经有text()或@href了，不需要再加
                    content_type = ""
                    if relativeXPath.find("/@href") >= 0 or relativeXPath.find("/text()") >= 0 or relativeXPath.find(
                            "::text()") >= 0:
                        content_type = ""
                    elif p["nodeType"] == 2:
                        content_type = "//@href"
                    elif p["nodeType"] == 4:
                        content_type = "//@src"
                    elif p["contentType"] == 1:
                        content_type = "/text()"
                    elif p["contentType"] == 0:
                        content_type = "//text()"
                    xpath = relativeXPath + content_type
                    if p["relative"]:
                        # if relativeXPath == "":
                        #     content = [loopElementHTML]
                        # else:
                        # 如果字串里有//即子孙查找，则不动语句
                        if relativeXPath.find("//") >= 0:
                            if xpath.startswith("/"):
                                full_path = "(" + parentPath + ")" + \
                                            "[" + str(index + 1) + "]" + \
                                            relativeXPath + content_type
                            else:  # 如果是id()这种形式，不需要包parentPath
                                full_path = xpath
                            try:
                                content = pageHTML.xpath(full_path)
                            except:
                                content = []
                        # 如果是id()这种形式，不需要包/html/body
                        elif not relativeXPath.startswith("/"):
                            try:
                                content = loopElementHTML.xpath(xpath)
                            except:
                                content = []
                        else:
                            content = loopElementHTML.xpath(
                                "/html/body/" + loopElementHTML[0][0].tag + xpath)
                    else:
                        # 如果是id()或(//div)[1]这种形式，不需要包/html/body
                        if xpath.find("/body") < 0 and xpath.startswith("/"):
                            xpath = "/html/body" + xpath
                        content = pageHTML.xpath(xpath)
                    if len(content) > 0:
                        # html = etree.tostring(content[0], encoding='utf-8').decode('utf-8')
                        # 拼接所有文本内容并去掉两边的空白
                        content = ' '.join(result.strip()
                                           for result in content if result.strip())
                        if p["nodeType"] == 2 or p["nodeType"] == 4:
                            base_url = self.browser.current_url
                            # 合并链接相对路径为绝对路径
                            content = urljoin(base_url, content)
                    else:
                        content = p["default"]
                        if not self.dataNotFoundKeys[p["name"]]:
                            self.print_and_log(
                                'Element %s not found with parameter name %s when extracting data, use default, this error will only show once' % (
                                    relativeXPath, p["name"]))
                            self.print_and_log(
                                "提取数据操作时，字段名 %s 对应XPath %s 未找到，使用默认值，本字段将不再重复报错" % (
                                    p["name"], relativeXPath))
                            self.dataNotFoundKeys[p["name"]] = True
                except Exception as e:
                    if not self.dataNotFoundKeys[p["name"]]:
                        self.print_and_log(
                            'Element %s not found with parameter name %s when extracting data, use default, this error will only show once' % (
                                relativeXPath, p["name"]))
                        self.print_and_log(
                            "提取数据操作时，字段名 %s 对应XPath %s 未找到（请查看原因，如是否翻页太快页面元素未加载出来），使用默认值，本字段将不再重复报错" % (
                                p["name"], relativeXPath))
                        self.dataNotFoundKeys[p["name"]] = True
                try:
                    self.outputParameters[p["name"]] = content
                except:
                    self.outputParameters[p["name"]] = p["default"]
        # 对于不能优化的操作，使用selenium执行
        for p in param["params"]:
            if not p["optimizable"]:
                content = ""
                relativeXPath = replace_field_values(
                    p["relativeXPath"], self.outputParameters, self)
                if not (p["contentType"] == 5 or p["contentType"] == 6):  # 如果不是页面标题或URL，去找元素
                    try:
                        # relativeXPath = relativeXPath.lower()
                        # relativeXPath = lowercase_tags_in_xpath(relativeXPath)
                        if p["relative"]:  # 是否相对xpath
                            if relativeXPath == "":  # 相对xpath有时候就是元素本身，不需要二次查找
                                element = loopElement
                            else:
                                # 如果字串里有//即子孙查找，则不动语句
                                if relativeXPath.find("//") >= 0:
                                    # full_path = "(" + parentPath + \
                                    #     relativeXPath + ")" + \
                                    #     "[" + str(index + 1) + "]"
                                    full_path = "(" + parentPath + ")" + \
                                                "[" + str(index + 1) + "]" + \
                                                relativeXPath
                                    element = self.browser.find_element(
                                        By.XPATH, full_path, iframe=p["iframe"])
                                else:
                                    element = loopElement.find_element(By.XPATH,
                                                                       relativeXPath[1:])
                        else:
                            element = self.browser.find_element(
                                By.XPATH, relativeXPath, iframe=p["iframe"])
                    except (
                    NoSuchElementException, InvalidSelectorException, StaleElementReferenceException) as e:  # 找不到元素的时候，使用默认值
                        # self.print_and_log(p)
                        try:
                            content = p["default"]
                        except Exception as e:
                            content = ""
                        self.outputParameters[p["name"]] = content
                        try:
                            if not self.dataNotFoundKeys[p["name"]]:
                                self.print_and_log(
                                    'Element %s not found with parameter name %s when extracting data, use default, this error will only show once' % (
                                        relativeXPath, p["name"]))
                                self.print_and_log(
                                    "提取数据操作时，字段名 %s 对应XPath %s 未找到，使用默认值，本字段将不再重复报错" % (
                                        p["name"], relativeXPath))
                                self.dataNotFoundKeys[p["name"]] = True
                        except:
                            pass
                        continue
                    except TimeoutException:  # 超时的时候设置超时值
                        self.print_and_log(
                            'Time out after set seconds when getting data')
                        try:
                            self.browser.execute_script('window.stop()')
                        except:
                            pass
                        if p["relative"]:  # 是否相对xpath
                            if relativeXPath == "":  # 相对xpath有时候就是元素本身，不需要二次查找
                                element = loopElement
                            else:
                                element = loopElement.find_element(By.XPATH,
                                                                   relativeXPath[1:])
                        else:
                            element = self.browser.find_element(
                                By.XPATH, relativeXPath, iframe=p["iframe"])
                        # rt.end()
                else:
                    element = self.browser.find_element(
                        By.XPATH, "//body", iframe=p["iframe"])
                try:
                    self.execute_code(
                        2, p["beforeJS"], p["beforeJSWaitTime"], element, iframe=p["iframe"])  # 执行前置js
                    content = self.get_content(p, element)
                except StaleElementReferenceException:  # 发生找不到元素的异常后，等待几秒重新查找
                    self.recordLog(
                        'StaleElementReferenceException: ' + relativeXPath)
                    time.sleep(3)
                    try:
                        if p["relative"]:  # 是否相对xpath
                            if relativeXPath == "":  # 相对xpath有时候就是元素本身，不需要二次查找
                                element = loopElement
                                self.recordLog(
                                    'StaleElementReferenceException: loopElement')
                            else:
                                element = loopElement.find_element(By.XPATH,
                                                                   relativeXPath[1:])
                                self.recordLog(
                                    'StaleElementReferenceException: loopElement+relativeXPath')
                        else:
                            element = self.browser.find_element(
                                By.XPATH, relativeXPath, iframe=p["iframe"])
                            self.recordLog(
                                'StaleElementReferenceException: relativeXPath')
                        content = self.get_content(p, element)
                    except StaleElementReferenceException:
                        self.recordLog(
                            'StaleElementReferenceException: ' + relativeXPath)
                        continue  # 再出现类似问题直接跳过
                self.outputParameters[p["name"]] = content
                self.execute_code(
                    2, p["afterJS"], p["afterJSWaitTime"], element, iframe=p["iframe"])  # 执行后置JS
        if param["recordASField"] > 0 and param["newLine"]:
            line = new_line(self.outputParameters,
                            self.maxViewLength, self.outputParametersRecord)
            self.OUTPUT.append(line)


def start_ipc_server(shutdown_event_to_set, key):
    """在一个新线程中启动一个远程控制服务器。用于在 nodejs 父进程和 python 子进程之间进行通信。
    :param shutdown_event_to_set: 用于设置主程序中的全局关闭事件
    :param key: 用于验证远程控制请求的密钥
    """
    
    class ShutdownHandler(BaseHTTPRequestHandler):
        def do_GET(self):
            request_key = self.headers.get('Authorization')
            if request_key != key and request_key != "Bearer " + key:
                self.send_response(403)
                self.end_headers()
                self.wfile.write(b'Forbidden: Invalid key.')
                return

            if self.path == '/shutdown':
                print("IPC server received shutdown command.")
                # 设置主程序中的全局关闭事件
                shutdown_event_to_set.set()
                self.send_response(200)
                self.end_headers()
                self.wfile.write(b'Shutdown signal received.')
            else:
                self.send_response(404)
                self.end_headers()
                self.wfile.write(b'Not Found.')
        def log_message(self, format, *args):
                # 覆盖此方法以禁止向 stderr 打印日志
                return

    # 绑定到 localhost 和一个随机可用端口 (port 0)
    httpd = HTTPServer(("127.0.0.1", 0), ShutdownHandler)
    
    # 获取被分配的端口号
    ipc_port = httpd.socket.getsockname()[1]

     # 关键步骤：将端口号打印到 stdout，以便 Node.js 父进程捕获
    print(f"IPC_SERVER_PORT:{ipc_port}", flush=True)
    
    # 在当前线程中运行服务器，直到被关闭
    httpd.serve_forever()


if __name__ == '__main__':
    # 如果需要调试程序，请在命令行参数中加入--keyboard 0 来禁用键盘监听以提升调试速度
    # If you need to debug the program, please add --keyboard 0 in the command line parameters to disable keyboard listening to improve debugging speed
    commandline_config = {
        "ids": [0],
        "saved_file_name": "",
        "user_data": False,
        "config_folder": "",
        "config_file_name": "config.json",
        "read_type": "remote",
        "headless": False,
        "server_address": "http://localhost:8074",
        "remote_control": False, # 是否开启远程控制服务器。这个所谓的“远程”其实是指 electron 主程序的控制，无法从外部网络访问（因为端口随机）
        "remote_control_key": "", # 如果开启远程控制，那么初始化时需要填写此 key，保证安全性
        "keyboard": True,  # 是否监听键盘输入
        "pause_key": "p",  # 暂停键
        "version": "0.6.3",
        "docker_driver": "",
        "user_folder": "",
    }
    c = Config(commandline_config)
    remote_key = c.remote_control_key
    c.remote_control_key = "hidden for security"  # 清空远程控制密钥，防止被其他人使用
    print(c)
    c.remote_control_key = remote_key  # 恢复远程控制密钥
    options = webdriver.ChromeOptions()
    driver_path = "chromedriver.exe"
    print(sys.platform, platform.architecture())
    if not os.path.exists(os.getcwd() + "/Data"):
        os.mkdir(os.getcwd() + "/Data")
    if sys.platform == "darwin" and platform.architecture()[0] == "64bit":
        options.binary_location = "EasySpider.app/Contents/Resources/app/chrome_mac64.app/Contents/MacOS/Google Chrome"
        options.add_extension(
            "EasySpider.app/Contents/Resources/app/XPathHelper.crx")
        driver_path = "EasySpider.app/Contents/Resources/app/chromedriver_mac64"
        print(driver_path)
        if c.config_folder == "":
            c.config_folder = os.path.expanduser(
                "~/Library/Application Support/EasySpider/")
    # 在 linux 里，EasySpider 打包版默认是在 EasySpider 文件夹内执行的，没有 windows 上那种外层的 25kb 启动器
    # 因此需要特殊处理并判断此时 chrome, chromedriver 和扩展的路径。
    elif sys.platform == "linux" and platform.architecture()[0] == "64bit" and os.path.exists(os.path.join(os.getcwd(), "resources")):
        print("Finding chromedriver in EasySpider",
              os.getcwd())
        # 相对于下一条检查语句，这里去掉了 EasySpider 文件夹这一层
        options.binary_location = "resources/app/chrome_linux64/chrome"
        driver_path = "resources/app/chrome_linux64/chromedriver_linux64"
        options.add_extension("resources/app/XPathHelper.crx")
    elif os.path.exists(os.getcwd() + "/EasySpider/resources"):  # 打包后的路径
        print("Finding chromedriver in EasySpider",
              os.getcwd() + "/EasySpider")
        if sys.platform == "win32" and platform.architecture()[0] == "32bit":
            options.binary_location = os.path.join(
                os.getcwd(), "EasySpider/resources/app/chrome_win32/chrome.exe")  # 指定chrome位置
            driver_path = os.path.join(
                os.getcwd(), "EasySpider/resources/app/chrome_win32/chromedriver_win32.exe")
            options.add_extension("EasySpider/resources/app/XPathHelper.crx")
        elif sys.platform == "win32" and platform.architecture()[0] == "64bit":
            options.binary_location = os.path.join(
                os.getcwd(), "EasySpider/resources/app/chrome_win64/chrome.exe")
            driver_path = os.path.join(
                os.getcwd(), "EasySpider/resources/app/chrome_win64/chromedriver_win64.exe")
            options.add_extension("EasySpider/resources/app/XPathHelper.crx")
        elif sys.platform == "linux" and platform.architecture()[0] == "64bit":
            options.binary_location = "EasySpider/resources/app/chrome_linux64/chrome"
            driver_path = "EasySpider/resources/app/chrome_linux64/chromedriver_linux64"
            options.add_extension("EasySpider/resources/app/XPathHelper.crx")
        else:
            print("Unsupported platform")
            sys.exit()
        print("Chrome location:", options.binary_location)
        print("Chromedriver location:", driver_path)
    elif os.path.exists(os.getcwd() + "/../ElectronJS"):
        # 软件dev用
        print("Finding chromedriver in EasySpider",
              os.getcwd() + "/ElectronJS")
        if sys.platform == "win32" and platform.architecture()[0] == "32bit":
            options.binary_location = os.path.join(
                os.getcwd(), "EasySpider/resources/app/chrome_win32/chrome.exe")  # 指定chrome位置
            driver_path = os.path.join(
                os.getcwd(), "EasySpider/resources/app/chrome_win32/chromedriver_win32.exe")
            options.add_extension("EasySpider/resources/app/XPathHelper.crx")
        elif sys.platform == "win32" and platform.architecture()[0] == "64bit":
            options.binary_location = "../ElectronJS/chrome_win64/chrome.exe"  # 指定chrome位置
            driver_path = "../ElectronJS/chrome_win64/chromedriver_win64.exe"
            options.add_extension("../ElectronJS/XPathHelper.crx")
        elif sys.platform == "linux" and platform.architecture()[0] == "64bit":
            options.binary_location = "../ElectronJS/chrome_linux64/chrome"
            driver_path = "../ElectronJS/chrome_linux64/chromedriver_linux64"
            options.add_extension("../ElectronJS/XPathHelper.crx")
        else:
            print("Unsupported platform for automatic detection. You need to specify chrome executable path, chromedriver path in code.")
            sys.exit()
    else:
        options.binary_location = "./chrome.exe"  # 指定chrome位置
        driver_path = "./chromedriver.exe"
        options.add_extension("XPathHelper.crx")

    options.add_experimental_option(
        'excludeSwitches', ['enable-automation'])  # 以开发者模式


    # 总结：
    # 0. 带Cookie需要用userdatadir
    # 1. chrome_options才是配置用户文件和chrome文件地址的正确选项
    # 2. User Profile文件夹的路径是：C:\Users\用户名\AppData\Local\Google\Chrome\User Data不要加Default
    # 3. 就算User Profile相同，chrome版本不同所存储的cookie信息也不同，也不能爬
    # 4. TMALL如果一直弹出验证码，而且无法通过验证，那么需要在其他浏览器上用
    try:
        with open(c.config_folder + c.config_file_name, "r", encoding='utf-8') as f:
            config = json.load(f)
            print("Config file path: " +
                  c.config_folder + c.config_file_name)
            absolute_user_data_folder = config["absolute_user_data_folder"]
    except:
        pass

    options.add_argument(
        "--disable-blink-features=AutomationControlled")  # TMALL 反扒
    # 阻止http -> https的重定向
    options.add_argument("--disable-features=CrossSiteDocumentBlockingIfIsolating,CrossSiteDocumentBlockingAlways,IsolateOrigins,site-per-process")
    options.add_argument("--disable-web-security")  # 禁用同源策略
    options.add_argument('-ignore-certificate-errors')
    options.add_argument('-ignore -ssl-errors')

    if c.headless:
        print("Headless mode")
        print("无头模式")
        options.add_argument("--headless")

    tmp_options = []
    for id in c.ids:
        tmp_options.append({"options": copy.deepcopy(options), "tmp_user_data_folder": ""})
    
    ipc_thread = None
    shutdown_event = Event()
    if c.remote_control:
        if c.remote_control_key == "":
            print("Remote control is enabled, but no remote control key is set, please set the --remote_control_key parameter to a non-empty value.")
            print("远程控制已启用，但未设置远程控制密钥，请将--remote_control_key参数设置为非空值。")
            sys.exit(1)
        else:
            ipc_thread = threading.Thread(target=start_ipc_server, args=(shutdown_event, c.remote_control_key), daemon=True)
            ipc_thread.start() 

    if c.user_data:
        tmp_user_folder_parent = os.path.join(os.getcwd(), "TempUserDataFolder")
        if not os.path.exists(tmp_user_folder_parent):
            os.mkdir(tmp_user_folder_parent)
        characters = string.ascii_letters + string.digits
        for i in range(len(c.ids)):
            options = tmp_options[i]["options"]
            options.add_argument("--profile-directory=Default")
            if c.user_folder == "":
                id = c.ids[i]
                # 从字符集中随机选择字符构成字符串
                random_string = ''.join(random.choice(characters) for i in range(10))
                tmp_user_data_folder = os.path.join(tmp_user_folder_parent, "user_data_" + str(id) + "_" + str(time.time()).replace(".","") + "_" + random_string)
                tmp_options[i]["tmp_user_data_folder"] = tmp_user_data_folder
                if os.path.exists(tmp_user_data_folder):
                    try:
                        shutil.rmtree(tmp_user_data_folder)
                    except:
                        pass
                print(f"Copying user data folder to: {tmp_user_data_folder}, please wait...")
                print(f"正在复制用户信息目录到: {tmp_user_data_folder}，请稍等...")
                if os.path.exists(absolute_user_data_folder):
                    try:
                        shutil.copytree(absolute_user_data_folder, tmp_user_data_folder)
                        print("User data folder copied successfully, if you exit the program before it finishes, please delete the temporary user data folder manually.")
                        print("用户信息目录复制成功，如果程序在运行过程中被手动退出，请手动删除临时用户信息目录。")
                    except:
                        tmp_user_data_folder = absolute_user_data_folder
                        print("Copy user data folder failed, use the original folder.")
                        print("复制用户信息目录失败，使用原始目录。")
                else:
                    tmp_user_data_folder = absolute_user_data_folder
                    print("Cannot find user data folder, create a new folder.")
                    print("未找到用户信息目录，创建新目录。")
                options.add_argument(
                    f'--user-data-dir={tmp_user_data_folder}')  # TMALL 反扒
                print(f"Use local user data folder: {tmp_user_data_folder}")
                print(f"使用本地用户信息目录: {tmp_user_data_folder}")
            else:
                options.add_argument(
                    f'--user-data-dir={c.user_folder}')
                print(f"Use specifed user data folder: {c.user_folder}, please note if you are using docker, this user folder path should be the path inside the docker container.")
                print(f"使用指定的用户信息目录: {c.user_folder}，请注意如果您正在使用docker，此用户文件夹路径应是容器内的路径。")
    print(
        "如果报错Selenium.common.exceptions.WebDriverException: Message: unknown error: Chrome failed to start: exited abnormally，说明有之前运行的Chrome实例没有正常关闭，请关闭之前打开的所有Chrome实例后再运行程序即可。")
    print(
        "If you get an error Selenium.common.exceptions.WebDriverException: Message: unknown error: Chrome failed to start: exited abnormally, it means that there is a Chrome instance that was not closed properly before, please close all Chrome instances that were opened before running the program.")

    threads = []
    for i in range(len(c.ids)):
        id = c.ids[i]
        options = tmp_options[i]["options"]
        print("id: ", id)
        if c.read_type == "remote":
            print("remote")
            try:
                content = requests.get(
                c.server_address + "/queryExecutionInstance?id=" + str(id))
                service = json.loads(content.text)  # 加载服务信息
            except:
                print("Cannot connect to the server, please make sure that the EasySpider Main Program is running, or you can change the --read_type parameter to 'local' to read the task information from the local task file without keeping the EasySpider Main Program running.")
                print("无法连接到服务器，请确保EasySpider主程序正在运行，或者您可以将--read_type参数更改为'local'，以实现从本地任务文件中读取任务信息而无需保持EasySpider主程序运行。")
        else:
            print("local")
            local_folder = os.path.join(os.getcwd(), "execution_instances")
            if sys.platform == "darwin":
                user_folder = os.path.expanduser(
                "~/Library/Application Support/EasySpider/")
                local_folder = os.path.join(user_folder, "execution_instances")
            file_path = os.path.join(local_folder, str(id) + ".json")
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                service = json.loads(content)  # 加载服务信息
        try:
            print("Task Name:", service["name"])
            print("任务名称:", service["name"])
        except:
            print(f"Cannot find task with id: {str(id)}, please check whether {str(id)}.json exists in the 'execution_instances' folder.")
            print(f"未找到id为{str(id)}的任务，请检查'execution_instances'文件夹中是否存在{str(id)}.json文件。")
            continue
        try:
            cloudflare = service["cloudflare"]
        except:
            cloudflare = 0
        if cloudflare == 0:
            options.add_argument('log-level=3')  # 隐藏日志
            path = os.path.join(os.path.abspath("./"), "Data", "Task_" + str(id), "files")
            print("文件下载路径|File Download path:", path)
            options.add_experimental_option("prefs", {
                # 设置文件下载路径
                "download.default_directory": path,
                "download.prompt_for_download": False,  # 禁止下载提示框
                "plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}],
                "download.directory_upgrade": True,
                "download.extensions_to_open": "applications/pdf",
                "plugins.always_open_pdf_externally": True,  # 总是在外部程序中打开PDF
                "safebrowsing_for_trusted_sources_enabled": False,
                "safebrowsing.enabled": False,
                'safebrowsing.disable_download_protection': True,
                'profile.default_content_settings.popups': 0,
            })
            try:
                if service["environment"] == 1:
                    options.add_experimental_option(
                        'mobileEmulation', {'deviceName': 'iPhone X'})  # 模拟iPhone X浏览
            except:
                pass
            try:
                browser = service["browser"]
            except:
                browser = "chrome"
            if browser == "chrome":
                if c.docker_driver == "":
                    print("Using local driver")
                    selenium_service = Service(executable_path=driver_path)
                    browser_t = MyChrome(service=selenium_service, options=options, mode='local_driver')
                else:
                    print("Using remote driver")
                    # Use docker driver, default address is http://localhost:4444/wd/hub
                    # Headless mode
                    # options.add_argument("--headless")
                    # print("Headless mode")
                    options.binary_location = ""
                    options.extensions.clear()
                    browser_t = MyChrome(command_executor=c.docker_driver, options=options, mode='remote_driver')
            elif browser == "edge":
                from selenium.webdriver.edge.service import Service as EdgeService
                from selenium.webdriver.edge.options import Options as EdgeOptions
                from myChrome import MyEdge
                selenium_service = EdgeService(executable_path="msedgedriver.exe")
                options = EdgeOptions()
                options.use_chromium = True
                options.add_argument("--ie-mode")
                options.add_argument("ie.edgepath=msedge.exe")
                browser_t = MyEdge(service=selenium_service, options=options)
        elif cloudflare == 1:
            if sys.platform == "win32":
                options.binary_location = "C:\\Program Files\\Google\\Chrome Beta\\Application\\chrome.exe"  # 需要用自己的浏览器
                browser_t = MyUCChrome(
                    options=options, driver_executable_path=driver_path)
                links = list(filter(isnotnull, service["links"].split("\n")))
                # open page in new tab
                browser_t.execute_script(
                    'window.open("' + links[0] + '","_blank");')
                time.sleep(5)  # wait until page has loaded
                browser_t.switch_to.window(
                    browser_t.window_handles[1])  # switch to new tab
                # browser_t = uc.Chrome()
            else:
                print("Cloudflare模式只支持Windows x64平台。")
                print(
                    "Cloudflare Mode only support on Windows x64 platform.")
                sys.exit()
        event = Event()
        event.set()
        thread = BrowserThread(browser_t, id, service,
                               c.version, event, c.saved_file_name, config=config, option=tmp_options[i], shutdown_event=shutdown_event, commandline_config=c)
        print("Thread with task id: ", id, " is created")
        threads.append(thread)
        thread.start()
        # Set the pause operation
        # if sys.platform != "linux":
        #     time.sleep(3)
        #     Thread(target=check_pause, args=("p", event)).start()
        # else:
        time.sleep(3)
        if c.pause_key == "p":
            try:
                pause_key = service["pauseKey"]
            except:
                pause_key = "p"
        else:
            pause_key = c.pause_key
        press_time = {"duration": 0, "is_pressed": False, "pause_key": pause_key}
        print("\n\n----------------------------------")
        print(
            "正在运行任务，长按键盘" + pause_key + "键可暂停任务的执行以便手工操作浏览器如输入验证码；如果想恢复任务的执行，请再次长按" + pause_key + "键。")
        print(
            "Running task, long press '" + pause_key + "' to pause the task for manual operation of the browser such as entering the verification code; If you want to resume the execution of the task, please long press '" + pause_key + "' again.")
        print("----------------------------------\n\n")
        # if cloudflare:
        #     print("过Cloudflare验证模式有时候会不稳定，如果无法通过验证则需要隔几分钟重试一次，或者可以更换新的用户信息文件夹再执行任务。")
        #     print("Passing the Cloudflare verification mode is sometimes unstable. If the verification fails, you need to try again every few minutes, or you can change to a new user information folder and then execute the task.")
        # 使用监听器监听键盘输入
    listener = None
    try:
        from pynput.keyboard import Key, Listener
        if c.keyboard:
            listener = Listener(on_press=on_press_creator(press_time, event),
                          on_release=on_release_creator(event, press_time))
    except:
        pass
        # print("您的操作系统不支持暂停功能。")
        # print("Your operating system does not support the pause function.")

    try:
        while (any(thread.is_alive() for thread in threads)):
            for thread in threads:
                thread.join(0.1)
    except (KeyboardInterrupt, SystemExit):
        print("程序被手动终止，正在关闭浏览器...")
        print("The program is manually terminated, closing the browser...")
        if not shutdown_event.is_set():
            shutdown_event.set()
    finally:
        if listener is not None and listener.is_alive():
            listener.stop()
